"""수납 대상자 명단 xlsx 내보내기·일괄 업로드."""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook, load_workbook

PAYMENT_HEADERS = [
    "접수번호",
    "접수ID",
    "한글성명",
    "영문성명",
    "생년월일",
    "시험장",
    "급수",
    "사진심사",
    "수납상태",
]

PAYMENT_GUIDE = [
    "수정·삭제 금지 — 매칭 키",
    "수정·삭제 금지 — 매칭 키(보조)",
    "참고용",
    "참고용",
    "참고용(YYYYMMDD)",
    "참고용",
    "참고용(Ⅰ/Ⅱ/동시)",
    "참고용 — 수정하지 마세요",
    "미수납 / 수납완료 (또는 X / O)\n※ 사진심사가 '승인'인 건만 수납완료 반영",
]

_PHOTO_LABEL = {
    "pending": "미심사",
    "approved": "승인",
    "rejected": "반려",
}

_LEVEL_LABEL = {"I": "Ⅰ", "II": "Ⅱ"}

_PAID_VALUES = frozenset({"o", "수납완료", "완료", "y", "yes", "1", "paid", "p"})
_UNPAID_VALUES = frozenset({"x", "미수납", "미납", "n", "no", "0", "unpaid", ""})


def payment_export_filename(*, round_no: int | None) -> str:
    round_part = f"제{round_no}회 " if round_no else ""
    return f"{round_part}TOPIK 수납대상자 명단(미얀마).xlsx"


def _level_ui(exam_level: str | None) -> str:
    lv = (exam_level or "I").upper()
    return _LEVEL_LABEL.get(lv, lv)


def payment_row_dict(row: dict[str, Any]) -> list[Any]:
    photo = _PHOTO_LABEL.get(str(row.get("photo_review_status") or "pending"), "미심사")
    return [
        row.get("application_no") or "",
        row.get("id") or "",
        row.get("name_ko") or "",
        row.get("name_en") or "",
        row.get("birth_date") or "",
        row.get("venue_name") or "",
        _level_ui(row.get("exam_level")),
        photo,
        "미수납",
    ]


def build_payment_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "수납대상"
    ws.append(PAYMENT_GUIDE)
    ws.append(PAYMENT_HEADERS)
    for row in rows:
        ws.append(payment_row_dict(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_payment_status(raw: Any) -> str | None:
    """'paid' | 'unpaid' | None(인식 불가)."""
    s = str(raw or "").strip().lower()
    if s in _PAID_VALUES:
        return "paid"
    if s in _UNPAID_VALUES:
        return "unpaid"
    # 한글 원문도 허용
    orig = str(raw or "").strip()
    if orig in ("수납완료", "완료"):
        return "paid"
    if orig in ("미수납", "미납"):
        return "unpaid"
    return None


def _norm_key(v: Any) -> str:
    return re.sub(r"\s+", "", str(v or "").strip()).lower()


def _header_map(header_row: list[Any]) -> dict[str, int]:
    m: dict[str, int] = {}
    for i, h in enumerate(header_row):
        key = str(h or "").strip()
        if key:
            m[key] = i
    return m


def _cell(row: tuple[Any, ...] | list[Any], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return row[idx]


def parse_payment_xlsx(data: bytes) -> list[dict[str, Any]]:
    """xlsx 바이트 → [{application_id, application_no, payment_status}, ...]"""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = None
    headers: list[Any] = []
    for i, row in enumerate(rows):
        cells = list(row or ())
        labels = [str(c or "").strip() for c in cells]
        if "접수번호" in labels or "수납상태" in labels:
            header_idx = i
            headers = cells
            break
    if header_idx is None:
        return []

    col = _header_map(headers)
    app_no_i = col.get("접수번호")
    app_id_i = col.get("접수ID")
    pay_i = col.get("수납상태")
    if pay_i is None:
        return []

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        cells = list(row or ())
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        status = parse_payment_status(_cell(cells, pay_i))
        if status is None:
            continue
        out.append(
            {
                "application_id": _cell(cells, app_id_i),
                "application_no": _cell(cells, app_no_i),
                "payment_status": status,
                "name_ko": _cell(cells, col.get("한글성명")),
                "name_en": _cell(cells, col.get("영문성명")),
            }
        )
    return out


def match_application(
    item: dict[str, Any],
    *,
    by_id: dict[str, Any],
    by_no: dict[str, Any],
) -> Any | None:
    app_id = str(item.get("application_id") or "").strip()
    if app_id and app_id in by_id:
        return by_id[app_id]
    app_no = _norm_key(item.get("application_no"))
    if app_no and app_no in by_no:
        return by_no[app_no]
    return None
