#!/usr/bin/env python3
"""Dump authoritative roster/exam-number/photo xlsx forms for BO_FE conformance."""
import sys
from openpyxl import load_workbook

FILES = [
    "docs/고객사 자료/연명부 양식/연명부 양식.xlsx",
    "docs/고객사 자료/연명부 양식/수험번호 부여 안내.xlsx",
    "docs/고객사 자료/연명부 양식/연명부 및 사진제출 안내.xlsx",
    "docs/고객사 자료/연명부 양식/[서식7] (시행기관용) 한국어능력시험(PBT) 국외 지원자 연명부.xlsx",
]


def dump(path):
    print("=" * 100)
    print("FILE:", path)
    print("=" * 100)
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print("  ERROR:", e)
        return
    for ws in wb.worksheets:
        print(f"\n--- SHEET: {ws.title}  (dims={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})")
        max_r = min(ws.max_row, 40)
        max_c = min(ws.max_column, 20)
        for r in range(1, max_r + 1):
            cells = []
            for c in range(1, max_c + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    cells.append(f"[{r},{c}]={v!r}")
            if cells:
                print("  " + " | ".join(cells))
        # merged cells
        if ws.merged_cells.ranges:
            print("  MERGED:", [str(m) for m in ws.merged_cells.ranges][:30])


if __name__ == "__main__":
    for f in FILES:
        dump(f)
