"""
TOPIK Myanmar WBS 세부 개발 일정 — 4대 메뉴 구조(고객사 0519) + 0526 수정사항 반영
================================================================
- 대상 파일: wbs/(아이뱅크) TOPIK_Myanmar_WBS 세부 개발 일정_v1.1.xlsx
- 전체 기간: 2026-05-20(수) ~ 2026-07-10(금)  (오픈일 2026-07-10)
- 단위테스트 5BD · 통합테스트 5BD
- FO-BO 연계 메뉴는 병행

[0526 변경 반영 항목]
  ㆍ FO_ACCOUNT : SNS 간편가입(구글 OAuth 2.0) · 비밀번호 6개월 주기 · 탈퇴 시 접수 취소
  ㆍ BO_CONTENT : 댓글/대댓글 기능(비밀글 연동) 추가
  ㆍ BO_APPLY   : 수납 취소(환불)·수납 시 사진+기본정보 동시 확인 단일 메뉴
  ㆍ FO_APPLY   : 접수 취소 가능 시점 (오프라인 수납 전까지)
  ㆍ BO_EXAM    : 시험장 좌석배치도·책임자 입력 제외
  ㆍ 전 모듈     : 알림 이메일 전용(문자 발송 제외)
  ㆍ 수험번호 발급: 수납 마감(7.26) 후 영문 성명 알파벳순 일괄 부여

[설계 원칙]
  ㆍ IA, 메뉴구조도(v1.0) 의 PAGE NO. 단위로 1행씩 작성
    (8 FO 모듈 + 7 BO 모듈 = 15 모듈, 총 136 작업행 + 2 테스트행)
  ㆍ 모듈별로 디자인·개발·검수 일정 기입
  ㆍ R12 까지는 기존 산출물 행(요구사항/IA/기능정의서) 유지
  ㆍ R13~R148  작업 본문
  ㆍ R149       단위테스트(5BD)
  ㆍ R150       통합테스트(5BD)
  ㆍ 헤더 수식(M3 등)은 R13:R148 범위로 갱신

[FO 4대 메뉴]
  1. TOPIK 안내 (시험 개요·소개·문항·평가)
  2. TOPIK 규정 (유의·답안·응시료·신분증)
  3. TOPIK 접수 (접수방법·시험접수 4단계·접수확인·수험표출력)
  4. 게시판     (공지·환불·정정·문의·FAQ)

[BO 신규]
  ㆍ 관리자 계정 관리(다중 계정·동시 접속)
  ㆍ 관리자 처리 이력(아이디별 변경 이력)
  ㆍ 수험번호 13자리 부여(국가 3 / 지역 3 / 수준 1 / 시험장 2 / 응시자 4)
  ㆍ 엑셀 내보내기(국외 지원자 연명부 11컬럼)
  ㆍ 사진 일괄 다운로드(지역/시험장/수준별 폴더 압축)
  ㆍ 시험장 관리(국가·지역·시험장 코드 마스터)
  ㆍ 환불·정보정정신청 관리 / 문의게시판 관리
"""

from __future__ import annotations

import sys
from copy import copy
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO = Path(__file__).resolve().parents[1]
WBS_PATH = REPO / "wbs" / "(아이뱅크) TOPIK_Myanmar_WBS 세부 개발 일정_v1.1.xlsx"

sys.path.insert(0, str(REPO / "IA, 메뉴구조도" / "scripts"))
from build_topik_ia_menu import FO_ROWS, BO_ROWS  # noqa: E402

# ────────────────────────────────────────────────────────────────
# 컬럼 인덱스 (1-based)
#   B(2)  구분       C(3)  타입       D(4)  1depth   E(5)  2depth
#   F(6)  3depth     G(7)  4depth     H(8)  Spec     I(9)  비고
#   M(13) 디자인 담당자 / N(14) 진행률 / O(15) 시작일 / P(16) 완료
#   Q(17) 개발  담당자 / R(18) 진행률 / S(19) 시작일 / T(20) 완료
#   U(21) 검수  담당자 / V(22) 진행률 / W(23) 시작일 / X(24) 완료
# ────────────────────────────────────────────────────────────────
COL_GUBUN, COL_TYPE, COL_D1, COL_D2, COL_D3, COL_D4, COL_SPEC, COL_NOTE = 2, 3, 4, 5, 6, 7, 8, 9
COL_DESIGN_PIC, COL_DESIGN_PROG, COL_DESIGN_START, COL_DESIGN_END = 13, 14, 15, 16
COL_DEV_PIC,    COL_DEV_PROG,    COL_DEV_START,    COL_DEV_END    = 17, 18, 19, 20
COL_QA_PIC,     COL_QA_PROG,     COL_QA_START,     COL_QA_END     = 21, 22, 23, 24

# 한국 공휴일 (2026-05-20 ~ 2026-07-31 구간) — 해당 구간에는 평일 공휴일 없음
HOLIDAYS: set[date] = set()

DEFAULT_PIC = "조준형"


def D(y: int, m: int, d: int) -> date:
    return date(y, m, d)


def is_bd(d: date) -> bool:
    return d.weekday() < 5 and d not in HOLIDAYS


def next_bd(d: date) -> date:
    while not is_bd(d):
        d += timedelta(days=1)
    return d


# ────────────────────────────────────────────────────────────────
# 모듈별 일정 (디자인 시작/끝, 개발 시작/끝, 검수 시작/끝)
#   - 모두 영업일이어야 함
# ────────────────────────────────────────────────────────────────
SCHEDULE: dict[str, tuple[tuple[int, int], ...]] = {
    # 형식: (디자인_시작, 디자인_끝, 개발_시작, 개발_끝, 검수_시작, 검수_끝)
    # 오픈 2026-07-10 기준 전체 압축 + 고객사 0526 수정사항 반영

    # ────────── FO ──────────
    # FO 공통 (GNB·언어·모바일·로그인가드·푸터)
    "FO_COMMON":   ((5, 20), (5, 22), (6,  1), (6,  5), (6, 22), (6, 24)),
    # FO 홈
    "FO_HOME":     ((5, 25), (5, 27), (6,  1), (6,  5), (6, 22), (6, 24)),
    # FO TOPIK 안내 (시험 개요·소개·문항·평가)
    "FO_INFO":     ((5, 26), (5, 28), (6,  8), (6, 10), (6, 22), (6, 23)),
    # FO TOPIK 규정 (유의·답안·응시료·신분증) — 0526: 동시접수 응시료 개별수납 표기 반영
    "FO_RULES":    ((5, 26), (5, 28), (6,  8), (6, 10), (6, 22), (6, 23)),
    # FO TOPIK 접수 — 0526: 접수 취소 가능 시점(수납 전) 로직 반영
    "FO_APPLY":    ((6,  1), (6,  5), (6,  8), (6, 17), (6, 22), (6, 24)),
    # FO 게시판 (공지·환불·정정·문의·FAQ)
    "FO_BOARD":    ((6,  3), (6,  5), (6,  8), (6, 17), (6, 22), (6, 24)),
    # FO 계정 — 0526: SNS 간편가입(구글 OAuth)·비번 6개월 주기·탈퇴시 접수취소
    "FO_ACCOUNT":  ((5, 25), (5, 29), (6,  8), (6, 19), (6, 22), (6, 24)),
    # FO 외부 링크
    "FO_EXTERNAL": ((6,  1), (6,  1), (6,  8), (6,  8), (6, 22), (6, 22)),

    # ────────── BO ──────────
    # BO 인증·공통 레이아웃
    "BO_AUTH":     ((5, 20), (5, 22), (6,  1), (6,  5), (6, 22), (6, 24)),
    # BO 대시보드
    "BO_DASH":     ((6,  3), (6,  5), (6, 15), (6, 17), (6, 23), (6, 24)),
    # BO 접수 관리 — 0526: 수납취소(환불)·수납시 사진+기본정보 동시확인 단일메뉴
    "BO_APPLY":    ((6,  1), (6,  5), (6,  8), (6, 19), (6, 22), (6, 24)),
    # BO 시험 관리 — 0526: 시험장 좌석배치도·책임자 입력 제외로 범위 축소
    "BO_EXAM":     ((6,  3), (6,  5), (6, 15), (6, 17), (6, 23), (6, 24)),
    # BO 콘텐츠 — 0526: 환불·정정/문의 댓글·대댓글(비밀글 연동) 추가
    "BO_CONTENT":  ((6,  4), (6,  5), (6, 15), (6, 19), (6, 23), (6, 24)),
    # BO 회원·약관 — 0526: 탈퇴시 진행중 접수 자동취소
    "BO_MEMBER":   ((5, 27), (5, 29), (6,  8), (6, 12), (6, 22), (6, 24)),
    # BO 시스템 (관리자 계정·처리 이력·사이트 보기·로그아웃)
    "BO_SYSTEM":   ((5, 27), (5, 29), (6,  8), (6, 12), (6, 22), (6, 24)),
}

# 단위테스트(5BD): 2026-06-25(목) ~ 2026-07-01(수)
# 통합테스트(5BD): 2026-07-02(목) ~ 2026-07-08(수)
# 오픈:            2026-07-10(금)
UNIT_TEST  = (D(2026, 6, 25), D(2026, 7,  1))
INTEG_TEST = (D(2026, 7,  2), D(2026, 7,  8))


# ────────────────────────────────────────────────────────────────
# 타입 약어 → 한글 명칭
# ────────────────────────────────────────────────────────────────
TYPE_LABEL = {
    "P":  "페이지",
    "S":  "섹션",
    "C":  "컴포넌트",
    "LP": "레이어 팝업",
    "MP": "모달",
    "L":  "외부링크",
}


def _row_kind(page_no: str) -> str:
    return page_no.rsplit("_", 1)[-1] if page_no and "_" in page_no else "P"


# ────────────────────────────────────────────────────────────────
# IA → WBS 행 변환
#   FO_ROWS / BO_ROWS 컬럼 정의 (build_topik_ia_menu.py 참조)
#     FO: (PAGE NO., Role, 1Depth, 2Depth, 3Depth, 4Depth,
#          화면명, spec, 접근권한, 관리기능, 메뉴 주요 설명, REQ ID, 비고)
#     BO: (PAGE NO., Role, 1Depth, 2Depth, 3Depth, 4Depth,
#          화면명, spec, 메뉴 주요 설명, REQ ID, 비고)
# ────────────────────────────────────────────────────────────────
def _ia_to_wbs(rows: list[tuple], side: str, is_fo: bool) -> list[dict]:
    out: list[dict] = []
    for r in rows:
        page_no = r[0]
        d1 = r[2] or ""
        d2 = r[3] or ""
        d3 = r[4] or ""
        d4 = r[5] or ""
        name = r[6] or ""
        spec = r[7] or "1차"
        note = (r[12] if is_fo else r[10]) or ""
        kind = _row_kind(page_no)
        # 루트 페이지(P)에서 2depth가 비어있는 경우 화면명을 2depth로 노출 (예: "메인 · 홈")
        d2_final = d2 or (name if kind == "P" else "")
        out.append({
            "side":  side,           # FO / BO
            "type":  TYPE_LABEL.get(kind, "페이지"),
            "d1":    d1,
            "d2":    d2_final,
            "d3":    d3,
            "d4":    d4,
            "spec":  spec,
            "note":  note,
            "module": _module_for(side, d1, page_no),
            "page_no": page_no,
        })
    return out


def _module_for(side: str, d1: str, page_no: str) -> str:
    """1Depth + PAGE NO. → SCHEDULE 모듈 키."""
    if side == "FO":
        if d1 == "공통":            return "FO_COMMON"
        if d1 == "홈":              return "FO_HOME"
        if d1 == "TOPIK 안내":      return "FO_INFO"
        if d1 == "TOPIK 규정":      return "FO_RULES"
        if d1 == "TOPIK 접수":      return "FO_APPLY"
        if d1 == "게시판":          return "FO_BOARD"
        if d1 == "계정":            return "FO_ACCOUNT"
        if d1 == "외부":            return "FO_EXTERNAL"
    else:
        if d1 in ("인증", "공통 레이아웃"):  return "BO_AUTH"
        if d1 == "대시보드":                return "BO_DASH"
        if d1 == "접수 관리":               return "BO_APPLY"
        if d1 == "시험 관리":               return "BO_EXAM"
        if d1 == "콘텐츠":                  return "BO_CONTENT"
        if d1 == "회원·약관":               return "BO_MEMBER"
        if d1 == "시스템":                  return "BO_SYSTEM"
    raise KeyError(f"매핑 누락: side={side} d1={d1} page_no={page_no}")


# ────────────────────────────────────────────────────────────────
# 엑셀 출력
# ────────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="C5CDD9")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_TOP    = Alignment(wrap_text=True, vertical="top",    horizontal="left")
ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def set_date(ws, row: int, col: int, value: date) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.number_format = "yyyy-mm-dd"
    cell.alignment = ALIGN_CENTER


def set_text(ws, row: int, col: int, value, align=None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = align or ALIGN_TOP


def _clear_data_area(ws, row_start: int, row_end: int) -> None:
    # 데이터 영역 안의 병합 셀 해제
    to_unmerge: list[str] = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= row_start and rng.max_row <= row_end:
            to_unmerge.append(str(rng))
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    for r in range(row_start, row_end + 1):
        for c in range(COL_GUBUN, COL_QA_END + 1):
            cell = ws.cell(r, c)
            if cell.coordinate in ws.merged_cells:
                continue
            try:
                cell.value = None
            except AttributeError:
                pass


def _validate_schedule() -> None:
    for key, ((sm, sd), (em, ed), (sm2, sd2), (em2, ed2), (sm3, sd3), (em3, ed3)) in SCHEDULE.items():
        days = [
            D(2026, sm, sd), D(2026, em, ed),
            D(2026, sm2, sd2), D(2026, em2, ed2),
            D(2026, sm3, sd3), D(2026, em3, ed3),
        ]
        assert all(is_bd(d) for d in days), f"{key}: 비영업일 포함"


def main() -> None:
    _validate_schedule()

    wb = load_workbook(WBS_PATH)
    ws = wb.active

    fo_rows = _ia_to_wbs(FO_ROWS, "FO", is_fo=True)
    bo_rows = _ia_to_wbs(BO_ROWS, "BO", is_fo=False)
    all_rows = fo_rows + bo_rows

    row_start = 13
    row_end   = row_start + len(all_rows) - 1
    unit_test_row  = row_end + 1
    integ_test_row = row_end + 2

    # 기존 데이터 영역 정리 (R13:R1062 → 충분히 큰 범위 청소)
    _clear_data_area(ws, row_start, max(ws.max_row, integ_test_row + 5))

    # 본문 작성
    prev_side = None
    prev_d1 = None
    for i, item in enumerate(all_rows):
        r = row_start + i

        side_label = item["side"] if item["side"] != prev_side else None
        prev_side = item["side"]
        d1_label = item["d1"] if (item["d1"] != prev_d1 or side_label) else None
        prev_d1 = item["d1"]

        set_text(ws, r, COL_GUBUN, side_label or "", ALIGN_CENTER)
        set_text(ws, r, COL_TYPE,  item["type"],     ALIGN_CENTER)
        set_text(ws, r, COL_D1,    d1_label or "",   ALIGN_TOP)
        set_text(ws, r, COL_D2,    item["d2"],       ALIGN_TOP)
        set_text(ws, r, COL_D3,    item["d3"],       ALIGN_TOP)
        set_text(ws, r, COL_D4,    item["d4"],       ALIGN_TOP)
        set_text(ws, r, COL_SPEC,  item["spec"],     ALIGN_CENTER)
        set_text(ws, r, COL_NOTE,  item["note"],     ALIGN_TOP)

        sched = SCHEDULE[item["module"]]
        ds = D(2026, *sched[0]); de = D(2026, *sched[1])
        vs = D(2026, *sched[2]); ve = D(2026, *sched[3])
        qs = D(2026, *sched[4]); qe = D(2026, *sched[5])

        set_text(ws, r, COL_DESIGN_PIC,  DEFAULT_PIC, ALIGN_CENTER)
        ws.cell(r, COL_DESIGN_PROG).value = 0
        ws.cell(r, COL_DESIGN_PROG).number_format = "0%"
        ws.cell(r, COL_DESIGN_PROG).alignment = ALIGN_CENTER
        set_date(ws, r, COL_DESIGN_START, ds)
        set_date(ws, r, COL_DESIGN_END,   de)

        set_text(ws, r, COL_DEV_PIC,     DEFAULT_PIC, ALIGN_CENTER)
        ws.cell(r, COL_DEV_PROG).value = 0
        ws.cell(r, COL_DEV_PROG).number_format = "0%"
        ws.cell(r, COL_DEV_PROG).alignment = ALIGN_CENTER
        set_date(ws, r, COL_DEV_START,    vs)
        set_date(ws, r, COL_DEV_END,      ve)

        set_text(ws, r, COL_QA_PIC,      DEFAULT_PIC, ALIGN_CENTER)
        ws.cell(r, COL_QA_PROG).value = 0
        ws.cell(r, COL_QA_PROG).number_format = "0%"
        ws.cell(r, COL_QA_PROG).alignment = ALIGN_CENTER
        set_date(ws, r, COL_QA_START,     qs)
        set_date(ws, r, COL_QA_END,       qe)

    # 단위테스트 / 통합테스트 행
    for r, label, (qs, qe) in (
        (unit_test_row,  "단위테스트", UNIT_TEST),
        (integ_test_row, "통합테스트", INTEG_TEST),
    ):
        set_text(ws, r, COL_GUBUN, label, ALIGN_CENTER)
        set_text(ws, r, COL_TYPE,  "테스트", ALIGN_CENTER)
        set_text(ws, r, COL_SPEC,  "1차",    ALIGN_CENTER)
        set_text(ws, r, COL_DEV_PIC, DEFAULT_PIC, ALIGN_CENTER) if label == "단위테스트" else None
        if label == "단위테스트":
            ws.cell(r, COL_DEV_PROG).value = 0
            ws.cell(r, COL_DEV_PROG).number_format = "0%"
            ws.cell(r, COL_DEV_PROG).alignment = ALIGN_CENTER
            set_date(ws, r, COL_DEV_START, qs)
            set_date(ws, r, COL_DEV_END,   qe)
        set_text(ws, r, COL_QA_PIC, DEFAULT_PIC, ALIGN_CENTER)
        ws.cell(r, COL_QA_PROG).value = 0
        ws.cell(r, COL_QA_PROG).number_format = "0%"
        ws.cell(r, COL_QA_PROG).alignment = ALIGN_CENTER
        set_date(ws, r, COL_QA_START, qs)
        set_date(ws, r, COL_QA_END,   qe)

    # 헤더 수식(M3 등) — R13:R{row_end} 범위로 업데이트
    ws.cell(3, 13).value = f"=COUNTA(N13:N{row_end})"
    ws.cell(4, 13).value = "=M3-M5"
    ws.cell(5, 13).value = f'=COUNTIF(N13:N{row_end}, "100%")'
    ws.cell(6, 13).value = "=IFERROR(M5/M3, 0)"

    ws.cell(3, 17).value = f"=COUNTA(R13:R{row_end})"
    ws.cell(4, 17).value = "=Q3-Q5"
    ws.cell(5, 17).value = f'=COUNTIF(R13:R{row_end}, "100%")'
    ws.cell(6, 17).value = "=IFERROR(Q5/Q3, 0)"

    ws.cell(3, 21).value = f"=COUNTA(V13:V{row_end})"
    ws.cell(4, 21).value = "=U3-U5"
    ws.cell(5, 21).value = f'=COUNTIF(V13:V{row_end}, "100%")'
    ws.cell(6, 21).value = "=IFERROR(U5/U3, 0)"

    wb.save(WBS_PATH)

    print("✓ WBS 일정 기입 완료 (고객사 0526 수정사항 반영)")
    print(f"  · FO 작업행:        {len(fo_rows)}  (R{row_start}~R{row_start + len(fo_rows) - 1})")
    print(f"  · BO 작업행:        {len(bo_rows)}  (R{row_start + len(fo_rows)}~R{row_end})")
    print(f"  · 단위테스트:        R{unit_test_row}  ({UNIT_TEST[0]} ~ {UNIT_TEST[1]}, 5BD)")
    print(f"  · 통합테스트:        R{integ_test_row}  ({INTEG_TEST[0]} ~ {INTEG_TEST[1]}, 5BD)")
    print(f"  · 최종 오픈:         2026-07-10(금)")


if __name__ == "__main__":
    main()
