"""
IA·기능정의서(v1.1) 기준 WBS 구조 동기화 — 일정·담당자 등 기존 값 유지
================================================================
- IA 소스: IA, 메뉴구조도/FO|BO/*_v1.1.xlsx (FO_IA / BO_IA 시트)
- 대상: wbs/(아이뱅크) TOPIK_Myanmar_WBS 세부 개발 일정_v1.1.xlsx
- R13~R152  작업 본문 (140행)
- R153      단위테스트
- R154      통합테스트
- 구조(B~I)만 IA 기준 갱신, 일정(M~X)은 키 매칭으로 보존
"""
from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

REPO = Path(__file__).resolve().parents[1]
WBS_PATH = REPO / "wbs" / "(아이뱅크) TOPIK_Myanmar_WBS 세부 개발 일정_v1.1.xlsx"
FO_IA_PATH = REPO / "IA, 메뉴구조도" / "FO" / "(아이뱅크) TOPIK_Myanmar_FO_IA, 메뉴구조도_v1.1.xlsx"
BO_IA_PATH = REPO / "IA, 메뉴구조도" / "BO" / "(아이뱅크) TOPIK_Myanmar_BO_IA, 메뉴구조도_v1.1.xlsx"

COL_GUBUN, COL_TYPE, COL_D1, COL_D2, COL_D3, COL_D4, COL_SPEC, COL_NOTE = 2, 3, 4, 5, 6, 7, 8, 9
COL_DESIGN_PIC, COL_DESIGN_PROG, COL_DESIGN_START, COL_DESIGN_END = 13, 14, 15, 16
COL_DEV_PIC, COL_DEV_PROG, COL_DEV_START, COL_DEV_END = 17, 18, 19, 20
COL_QA_PIC, COL_QA_PROG, COL_QA_START, COL_QA_END = 21, 22, 23, 24
SCHED_COLS = range(COL_DESIGN_PIC, COL_QA_END + 1)

TYPE_LABEL = {
    "P": "페이지", "S": "섹션", "C": "컴포넌트",
    "LP": "레이어 팝업", "MP": "모달", "L": "외부링크",
}

ALIGN_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

# 오픈 2026-07-10(금) 기준 — 단위테스트(5BD) → 통합테스트(5BD) → 오픈
UNIT_TEST  = (date(2026, 6, 25), date(2026, 7,  1))   # 6/25(목) ~ 7/01(수)
INTEG_TEST = (date(2026, 7,  2), date(2026, 7,  8))   # 7/02(목) ~ 7/08(수)


def _row_kind(page_no: str) -> str:
    return str(page_no).rsplit("_", 1)[-1] if page_no and "_" in str(page_no) else "P"


def _norm(val) -> str:
    return " ".join(str(val or "").split())


def _read_ia_rows(path: Path, is_fo: bool) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out: list[dict] = []
    for r in range(6, ws.max_row + 1):
        page_no = ws.cell(r, 2).value
        if not page_no:
            continue
        d1 = _norm(ws.cell(r, 4).value)
        d2 = _norm(ws.cell(r, 5).value)
        d3 = _norm(ws.cell(r, 6).value)
        d4 = _norm(ws.cell(r, 7).value)
        name = _norm(ws.cell(r, 8).value)
        spec = ws.cell(r, 9).value or "1차"
        note_col = 14 if is_fo else 12
        note = ws.cell(r, note_col).value or ""
        kind = _row_kind(page_no)
        d2_final = d2 or (name if kind == "P" else "")
        out.append({
            "page_no": str(page_no),
            "side": "FO" if is_fo else "BO",
            "type": TYPE_LABEL.get(kind, "페이지"),
            "d1": d1,
            "d2": d2_final,
            "d3": d3,
            "d4": d4,
            "spec": spec,
            "note": note,
        })
    return out


def _content_key(item: dict) -> tuple:
    return (
        item["side"],
        item["type"],
        item["d1"],
        item["d2"],
        item["d3"],
        item["d4"],
    )


def _loose_key(item: dict) -> tuple:
    return (item["side"], item["d1"], item["d2"], item["d3"], item["type"])


def _module_key(side: str, d1: str) -> tuple:
    return (side, d1)


def _extract_schedule(ws, row: int) -> dict:
    sched: dict = {}
    for c in SCHED_COLS:
        sched[c] = {
            "value": ws.cell(row, c).value,
            "number_format": ws.cell(row, c).number_format,
            "alignment": copy(ws.cell(row, c).alignment),
            "font": copy(ws.cell(row, c).font),
            "fill": copy(ws.cell(row, c).fill),
            "border": copy(ws.cell(row, c).border),
        }
    return sched


def _read_existing_schedules(ws) -> tuple[dict, dict, dict]:
    """기존 WBS에서 일정 데이터 추출 (정확키 / 느슨한키 / 모듈키)."""
    exact: dict[tuple, dict] = {}
    loose: dict[tuple, dict] = {}
    module: dict[tuple, dict] = {}

    prev_side = prev_d1 = ""
    for r in range(13, ws.max_row + 1):
        gubun = ws.cell(r, COL_GUBUN).value or prev_side
        if ws.cell(r, COL_GUBUN).value:
            prev_side = gubun
        d1 = ws.cell(r, COL_D1).value or prev_d1
        if ws.cell(r, COL_D1).value:
            prev_d1 = d1
        typ = ws.cell(r, COL_TYPE).value
        if typ in (None, "테스트"):
            continue
        item = {
            "side": prev_side,
            "type": typ,
            "d1": _norm(d1),
            "d2": _norm(ws.cell(r, COL_D2).value),
            "d3": _norm(ws.cell(r, COL_D3).value),
            "d4": _norm(ws.cell(r, COL_D4).value),
        }
        sched = _extract_schedule(ws, r)
        ck = _content_key(item)
        if ck not in exact:
            exact[ck] = sched
        lk = _loose_key(item)
        if lk not in loose:
            loose[lk] = sched
        mk = _module_key(item["side"], item["d1"])
        module[mk] = sched
    return exact, loose, module


def _lookup_schedule(item: dict, exact, loose, module) -> dict | None:
    sched = exact.get(_content_key(item))
    if sched:
        return sched
    sched = loose.get(_loose_key(item))
    if sched:
        return sched
    return module.get(_module_key(item["side"], item["d1"]))


def _apply_schedule(ws, row: int, sched: dict, template_row: int) -> None:
    for c in SCHED_COLS:
        src = sched.get(c)
        cell = ws.cell(row, c)
        if src:
            cell.value = src["value"]
            cell.number_format = src["number_format"]
            cell.alignment = copy(src["alignment"])
            cell.font = copy(src["font"])
            cell.fill = copy(src["fill"])
            cell.border = copy(src["border"])
        else:
            tmpl = ws.cell(template_row, c)
            cell.value = tmpl.value
            cell.number_format = tmpl.number_format
            cell.alignment = copy(tmpl.alignment)


def _set_text(ws, row: int, col: int, value, align=None) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    cell.alignment = align or ALIGN_TOP


def _set_date(ws, row: int, col: int, value: date) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    cell.number_format = "yyyy-mm-dd"
    cell.alignment = ALIGN_CENTER


def _clear_note_merges(ws, row_start: int, row_end: int) -> None:
    to_unmerge = [
        str(rng)
        for rng in ws.merged_cells.ranges
        if rng.min_row >= row_start and rng.max_row <= row_end and rng.min_col == COL_NOTE
    ]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def _merge_note(ws, row: int) -> None:
    ws.merge_cells(
        start_row=row, start_column=COL_NOTE,
        end_row=row, end_column=COL_NOTE + 3,
    )


def main() -> None:
    ia_rows = _read_ia_rows(FO_IA_PATH, is_fo=True) + _read_ia_rows(BO_IA_PATH, is_fo=False)
    print(f"IA rows: {len(ia_rows)} (FO={sum(1 for x in ia_rows if x['side']=='FO')}, "
          f"BO={sum(1 for x in ia_rows if x['side']=='BO')})")

    wb = load_workbook(WBS_PATH)
    ws = wb.active

    exact, loose, module = _read_existing_schedules(ws)
    template_row = 13

    row_start = 13
    row_end = row_start + len(ia_rows) - 1
    unit_test_row = row_end + 1
    integ_test_row = row_end + 2

    _clear_note_merges(ws, row_start, max(ws.max_row, integ_test_row + 5))

    prev_side = prev_d1 = None
    matched = fallback = 0

    for i, item in enumerate(ia_rows):
        r = row_start + i
        side_label = item["side"] if item["side"] != prev_side else None
        prev_side = item["side"]
        d1_label = item["d1"] if (item["d1"] != prev_d1 or side_label) else None
        prev_d1 = item["d1"]

        _set_text(ws, r, COL_GUBUN, side_label or "", ALIGN_CENTER)
        _set_text(ws, r, COL_TYPE, item["type"], ALIGN_CENTER)
        _set_text(ws, r, COL_D1, d1_label or "", ALIGN_TOP)
        _set_text(ws, r, COL_D2, item["d2"], ALIGN_TOP)
        _set_text(ws, r, COL_D3, item["d3"], ALIGN_TOP)
        _set_text(ws, r, COL_D4, item["d4"], ALIGN_TOP)
        _set_text(ws, r, COL_SPEC, item["spec"], ALIGN_CENTER)
        _set_text(ws, r, COL_NOTE, item["note"], ALIGN_TOP)
        _merge_note(ws, r)

        sched = _lookup_schedule(item, exact, loose, module)
        if sched:
            _apply_schedule(ws, r, sched, template_row)
            matched += 1
        else:
            _apply_schedule(ws, r, module.get(_module_key(item["side"], item["d1"]), {}), template_row)
            fallback += 1

    # 단위/통합 테스트 행
    old_unit = _extract_schedule(ws, 149) if ws.cell(149, COL_TYPE).value == "테스트" else None
    old_integ = _extract_schedule(ws, 150) if ws.cell(150, COL_TYPE).value == "테스트" else None

    for r, label, (qs, qe), old in (
        (unit_test_row, "단위테스트", UNIT_TEST, old_unit),
        (integ_test_row, "통합테스트", INTEG_TEST, old_integ),
    ):
        for c in range(COL_GUBUN, COL_QA_END + 1):
            ws.cell(r, c).value = None
        _set_text(ws, r, COL_GUBUN, label, ALIGN_CENTER)
        _set_text(ws, r, COL_TYPE, "테스트", ALIGN_CENTER)
        _set_text(ws, r, COL_SPEC, "1차", ALIGN_CENTER)
        if old:
            _apply_schedule(ws, r, old, template_row)
        if label == "단위테스트":
            if not old or not old[COL_DEV_START]["value"]:
                _set_date(ws, r, COL_DEV_START, qs)
                _set_date(ws, r, COL_DEV_END, qe)
        if not old or not old[COL_QA_START]["value"]:
            _set_date(ws, r, COL_QA_START, qs)
            _set_date(ws, r, COL_QA_END, qe)

    # 이전 테스트 행 잔여 데이터 정리
    for r in range(integ_test_row + 1, integ_test_row + 6):
        for c in range(COL_GUBUN, COL_QA_END + 1):
            ws.cell(r, c).value = None

    # 헤더 수식 갱신
    ws.cell(3, 13).value = f"=COUNTA(N{row_start}:N{row_end})"
    ws.cell(4, 13).value = "=M3-M5"
    ws.cell(5, 13).value = f'=COUNTIF(N{row_start}:N{row_end}, "100%")'
    ws.cell(6, 13).value = "=IFERROR(M5/M3, 0)"

    ws.cell(3, 17).value = f"=COUNTA(R{row_start}:R{row_end})"
    ws.cell(4, 17).value = "=Q3-Q5"
    ws.cell(5, 17).value = f'=COUNTIF(R{row_start}:R{row_end}, "100%")'
    ws.cell(6, 17).value = "=IFERROR(Q5/Q3, 0)"

    ws.cell(3, 21).value = f"=COUNTA(V{row_start}:V{row_end})"
    ws.cell(4, 21).value = "=U3-U5"
    ws.cell(5, 21).value = f'=COUNTIF(V{row_start}:V{row_end}, "100%")'
    ws.cell(6, 21).value = "=IFERROR(U5/U3, 0)"

    wb.save(WBS_PATH)

    print("✓ WBS IA v1.1 동기화 완료")
    print(f"  · 작업행: R{row_start}~R{row_end} ({len(ia_rows)}행)")
    print(f"  · 단위테스트: R{unit_test_row}  · 통합테스트: R{integ_test_row}")
    print(f"  · 일정 보존: {matched}행 매칭, {fallback}행 모듈 fallback")


if __name__ == "__main__":
    main()
