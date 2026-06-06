#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기능정의서 xlsx 후처리:
- 일반인이 읽기 어려운 개발·영문 용어를 쉬운 한국어로 치환
- 메인 표의 '비고'는 논의·협의·운영 확정 등 이해관계자 합의가 필요한 내용만 남김
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FO_XLSX = ROOT / "FO" / "엑셀"
BO_XLSX = ROOT / "BO" / "엑셀"

TYPE_MAP = {
    "component": "화면 공통 부품",
    "section": "화면 구역",
    "section (예정)": "화면 구역(예정)",
    "page": "페이지",
    "page (예정)": "페이지(예정)",
    "function": "기능",
    "modal popup": "안내 창",
    "modal popup (예정)": "안내 창(예정)",
    "layer popup": "위쪽 안내 창",
    "logic": "동작 규칙",
}

# (pattern, replacement, flags) — 긴 문구·고유명사 우선
REGEX_REPLACEMENTS: list[tuple[str, str, int]] = [
    (r"iOS safe-area-inset", "아이폰 노치·하단 영역(안전 여백)", re.IGNORECASE),
    (r"safe-area-inset", "화면 안전 여백(노치 등)", re.IGNORECASE),
    (r"클라이언트\s*i18n", "다국어(브라우저에서 즉시 반영)", re.IGNORECASE),
    (r"서버 세션\s*\+\s*JWT\s*등", "서버 로그인 상태 + 전자 증표 등", 0),
    (r"\bJWT\b", "로그인 확인용 전자 증표", re.IGNORECASE),
    (r"\bCSRF\b", "위조 요청 차단", re.IGNORECASE),
    (r"\bBcrypt\b", "비밀번호 암호 저장 방식", re.IGNORECASE),
    (r"\bOAuth\b|\boauth\b", "외부 간편 로그인 규격", re.IGNORECASE),
    (r"\bWCAG\b", "웹 접근성 국제 지침", re.IGNORECASE),
    (r"\bSSR\b|\bssr\b", "서버에서 먼저 만든 페이지", re.IGNORECASE),
    (r"\bSEO\b|\bseo\b", "검색 노출", re.IGNORECASE),
    (r"\bCDN\b|\bcdn\b", "전 세계 속도 높이는 파일 배포망", re.IGNORECASE),
    (r"\bREST\b|\bRest\b|\brest\b", "웹 표준 통신 규격", 0),
    (r"\bJSON\b|\bjson\b", "데이터 교환 형식(JSON)", 0),
    (r"\bXHR\b|\bFetch\b|\bfetch\b", "웹 페이지의 서버 요청", 0),
    (r"\bHTML\b(?![가-힣])", "웹 페이지 문법(HTML)", 0),
    (r"\bcursor\b(?![가-힣])", "마우스 커서", re.IGNORECASE),
    (r"\bSMTP\b|\bsmtp\b", "이메일 발송 규격", re.IGNORECASE),
    (r"\bOTP\b|\botp\b", "일회용 인증번호", re.IGNORECASE),
    (r"\bTOTP\b|\btotp\b", "시간 기반 일회용 인증", re.IGNORECASE),
]

WORD_REPLACEMENTS: list[tuple[str, str]] = [
    ("i18n", "다국어 전환"),
    ("I18N", "다국어 전환"),
    ("localStorage", "브라우저 저장"),
    ("sessionStorage", "탭별 브라우저 저장"),
    ("sticky", "스크롤 시 위쪽 고정"),
    ("Sticky", "스크롤 시 위쪽 고정"),
    ("hover", "마우스를 올렸을 때"),
    ("Hover", "마우스를 올렸을 때"),
    ("tooltip", "짧은 설명 도움말"),
    ("modal", "안내 창"),
    ("Modal", "안내 창"),
    ("GNB", "상단 메뉴 바"),
    ("SPA", "한 페이지처럼 보이도록 전환되는 방식"),
    ("aria-", "접근성 속성 aria-"),
    ("ARIA", "접근성(보조기기용) 속성 묶음"),
    ("DOM", "화면 안의 요소"),
    ("dom", "화면 안의 요소"),
    ("API", "시스템 연동 프로그램"),
    ("Ux", "UX"),
    ("UX", "사용 경험"),
    ("UI", "화면 디자인·조작 방식"),
    ("slug", "주소 이름"),
    ("pixel", "화소"),
    ("px", "화소(px)"),
    ("fallback", "없을 때 대체값"),
    ("Fallback", "없을 때 대체값"),
    ("lazy load", "필요할 때만 불러오기"),
    ("Lazy Load", "필요할 때만 불러오기"),
    ("routing", "화면 이동 규칙"),
    ("Router", "화면 이동 처리기"),
    ("Cached", "임시 저장된"),
]

EDGE_WORD_RE = [
    (re.compile(r"\bfocus\b", re.IGNORECASE), "초점(키보드·스크린리더) 이동"),
    (re.compile(r"\bscroll(s|ing|ed)?\b", re.IGNORECASE), "스크롤"),
    (re.compile(r"\bcache(s|d)?\b", re.IGNORECASE), "임시 저장"),
    (re.compile(r"\btoken(s)?\b", re.IGNORECASE), "접속 증표"),
    (re.compile(r"\bhook(s)?\b", re.IGNORECASE), "이벤트 연결 처리"),
]

NOTE_TRIGGERS = (
    "협의",
    "논의",
    "합의",
    "확정",
    "미정",
    "별도 ",
    "정책",
    "통일",
    "동기화",
    "매핑",
    "검토",
    "향후",
    "추후",
    "권장",
    "운영 시",
    "실운영",
    "실제 운영",
    "도입 필수",
    "필수",
    "예정)",
    "(예정",
    "맞춤",
    "조율",
    "조정",
    "버전",
    "변경 시",
)


def simplify_cell_text(val: object) -> object:
    if val is None or not isinstance(val, str):
        return val
    s = val
    # 타입 고정 문자열 치환(전각·공백 허용)
    st = s.strip()
    if st in TYPE_MAP:
        return TYPE_MAP[st]
    for pat, rep, flags in REGEX_REPLACEMENTS:
        s = re.sub(pat, rep, s, flags=flags)
    for w, rep in WORD_REPLACEMENTS:
        if w in s:
            s = s.replace(w, rep)
    for cre, rep in EDGE_WORD_RE:
        s = cre.sub(rep, s)
    # 남은 짧은 잡영어 패턴 — 문맥상 자주 등장하는 것만
    s = re.sub(r"\bbox-shadow\b", "그림자(화면 디자인)", s, flags=re.IGNORECASE)
    s = re.sub(r"\bviewport\b", "브라우저 보이는 영역", s, flags=re.IGNORECASE)
    s = re.sub(r"\biframe\b", "다른 페이지 끌어오기 창", s, flags=re.IGNORECASE)
    s = re.sub(r"\bdict\b", "문자 사전", s)
    s = re.sub(r"\bplaceholder\b", "입력 안내 문구", s, flags=re.IGNORECASE)
    return s


def _garbage_note(s: str) -> bool:
    t = s.strip()
    if len(t) <= 2:
        return True
    junk = {"라벨", "링크", "입력", "ID", "표시 위치", "상세", "데이터"}
    if t in junk:
        return True
    if t.startswith("en>`") or "class=\"active\"" in t or "var(--" in t:
        return True
    return False


def consult_only_note(s: str) -> str:
    """비고: 합의·논의·운영 방침 결정이 필요한 문장만 유지."""
    if not s or not str(s).strip():
        return ""
    s = str(s).strip()
    s = simplify_cell_text(s)
    if not isinstance(s, str):
        return ""
    if _garbage_note(s):
        return ""
    if "참조" in s and not any(k in s for k in ("협의", "논의", "합의", "검토", "확인")):
        return ""
    if any(k in s for k in NOTE_TRIGGERS):
        return s.strip()
    # 운영·보안·외부 링크 정책 등(트리거 없어도 합의 필요로 보이는 짧은 문장)
    if "NIIED" in s or "운영" in s or "보안" in s or "개인정보" in s:
        return s.strip()
    return ""


def process_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        hdr = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        if not hdr or "비고" not in hdr:
            continue
        col_idx = {h: i + 1 for i, h in enumerate(hdr) if h}
        if "비고" not in col_idx:
            continue
        note_col = col_idx["비고"]
        text_cols = [
            c
            for c in [
                "Role",
                "1Depth",
                "2Depth",
                "3Depth",
                "4Depth",
                "타입",
                "화면명",
                "PC/MO",
                "spec",
                "접근 권한",
                "기능 설명",
            ]
            if c in col_idx
        ]
        for r in range(3, ws.max_row + 1):
            for cn in text_cols:
                v = ws.cell(r, col_idx[cn]).value
                nv = simplify_cell_text(v)
                if nv != v:
                    ws.cell(r, col_idx[cn], value=nv)
            nv = simplify_cell_text(ws.cell(r, note_col).value)
            nv = consult_only_note(str(nv or ""))
            ws.cell(r, note_col, value=nv)

    # 상세설명 시트 — 내용 열만
    if "상세설명" in wb.sheetnames:
        ws = wb["상세설명"]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 3).value
            nv = simplify_cell_text(v)
            if nv != v:
                ws.cell(r, 3, value=nv)

    for name in wb.sheetnames:
        if name.endswith("_상세"):
            ws = wb[name]
            hdr = [ws.cell(1, c).value for c in range(1, 5)]
            if hdr and hdr[2] == "내용":
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(r, 3).value
                    nv = simplify_cell_text(v)
                    if nv != v:
                        ws.cell(r, 3, value=nv)

    wb.save(path)


def main() -> None:
    for folder in (FO_XLSX, BO_XLSX):
        for p in sorted(folder.glob("*.xlsx")):
            process_workbook(p)
            print(f"OK {p.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
