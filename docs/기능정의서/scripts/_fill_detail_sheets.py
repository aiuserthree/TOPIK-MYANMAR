#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기능정의서 상세설명 시트 일괄 채움 스크립트.

이미 작성된 시트(예: FO 00, FO 01, BO 00, BO 01, BO 02)는 건드리지 않고,
비어 있는 상세설명 시트만 골라 5섹션(화면 흐름도/주요 검증 사항/연동 페이지 및 기능/
기술적 고려사항/참고 사항)을 일괄 작성한다.

대상:
  FO/02_TOPIK안내, FO/03_TOPIK규정, FO/04_TOPIK접수, FO/05_게시판,
  FO/06_계정_회원가입·로그인·내정보, FO/09_외부링크,
  BO/03_시험관리, BO/04_콘텐츠관리, BO/05_회원약관관리, BO/06_시스템_관리자계정·처리이력
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
FO_DIR = ROOT / "FO"
BO_DIR = ROOT / "BO"

THIN = Side(style="thin", color="C5CDD9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill(start_color="003478", end_color="003478", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
SECTION_FONT = Font(bold=True, color="003478", size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def write_detail(file_path: Path, content_blocks: list[tuple[str, list[str]]]) -> None:
    """비어 있는 상세설명 시트에 (섹션, 라인목록) 블록을 채운다."""
    wb = load_workbook(file_path)
    if "상세설명" in wb.sheetnames:
        ws = wb["상세설명"]
        # 기존 비어있는지 확인 (이미 채워진 시트는 보존)
        nonempty = 0
        for r in range(1, (ws.max_row or 0) + 1):
            for c in range(1, (ws.max_column or 0) + 1):
                if ws.cell(r, c).value:
                    nonempty += 1
                    break
        if nonempty >= 5:
            print(f"  [skip] {file_path.name} (이미 작성됨, {nonempty}행)")
            return
        del wb["상세설명"]

    ws = wb.create_sheet("상세설명")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 130

    h1 = ws.cell(1, 1, value="섹션")
    h2 = ws.cell(1, 2, value="내용")
    for c in (h1, h2):
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
    ws.row_dimensions[1].height = 22

    r = 2
    for section, lines in content_blocks:
        first = True
        for line in lines:
            if first:
                cell_a = ws.cell(r, 1, value=section)
                cell_a.font = SECTION_FONT
                cell_a.fill = SECTION_FILL
                cell_a.alignment = WRAP_TOP
                cell_a.border = BOX
                first = False
            cell_b = ws.cell(r, 2, value=line)
            cell_b.alignment = WRAP_TOP
            cell_b.border = BOX
            r += 1
        # 섹션 사이 한 줄 비움
        r += 1

    wb.save(file_path)
    print(f"  [OK]  {file_path.name} → {r - 2}행 작성")


# 콘텐츠 블록은 별도 모듈에서 가져옴 (가독성 + 점검 용이)
from _detail_blocks import BLOCKS  # noqa: E402


def main() -> None:
    print("═══════════════════════════════════════════════════════")
    print("FO 기능정의서 상세설명 시트 작성")
    print("═══════════════════════════════════════════════════════")
    fo_targets = [
        ("02_TOPIK안내_기능정의서.xlsx", "FO_02"),
        ("03_TOPIK규정_기능정의서.xlsx", "FO_03"),
        ("04_TOPIK접수_기능정의서.xlsx", "FO_04"),
        ("05_게시판_기능정의서.xlsx", "FO_05"),
        ("06_계정_회원가입·로그인·내정보_기능정의서.xlsx", "FO_06"),
        ("09_외부링크_기능정의서.xlsx", "FO_09"),
    ]
    for fname, key in fo_targets:
        fp = FO_DIR / fname
        if not fp.exists():
            print(f"  [skip] {fname} (파일 없음)")
            continue
        write_detail(fp, BLOCKS[key])

    print()
    print("═══════════════════════════════════════════════════════")
    print("BO 기능정의서 상세설명 시트 작성")
    print("═══════════════════════════════════════════════════════")
    bo_targets = [
        ("03_시험관리_기능정의서.xlsx", "BO_03"),
        ("04_콘텐츠관리_기능정의서.xlsx", "BO_04"),
        ("05_회원약관관리_기능정의서.xlsx", "BO_05"),
        ("06_시스템_관리자계정·처리이력_기능정의서.xlsx", "BO_06"),
    ]
    for fname, key in bo_targets:
        fp = BO_DIR / fname
        if not fp.exists():
            print(f"  [skip] {fname} (파일 없음)")
            continue
        write_detail(fp, BLOCKS[key])


if __name__ == "__main__":
    main()
