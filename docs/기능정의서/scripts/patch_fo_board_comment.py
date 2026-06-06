#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FO 게시판(05) 댓글/대댓글 기능 패치 — 고객사 0526 수정사항

대상 파일:
  FO/05_게시판_기능정의서.xlsx
  FO/_FO_기능정의서_통합_v1.1.xlsx

수정 항목:
  - 환불·정보정정신청 및 문의게시판 상세 페이지에 댓글/대댓글 기능 추가
  - 비밀글의 댓글/대댓글은 자동 비밀글 처리(작성자·관리자만 열람)
  - 알림 발송: 이메일 전용(문자 제외, 고객사 0526)
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT   = Path(__file__).resolve().parents[1]
FO_DIR = ROOT / "FO"

WRAP_TOP = Alignment(wrap_text=True, vertical="top")

PATCHES: dict[str, str] = {

    # ── 환불·정보정정신청 페이지 개요 ─────────────────────────────────────
    "TPKM_FO_5_2_0_0_0_P": (
        "**1. 페이지 개요**\n"
        "- '게시판 · 환불·정보정정신청'(`refund-correction.html`).\n"
        "- 고객사 수정(0519) 신규 — 자유게시판 형식 + 비밀글 옵션, 로그인 필수.\n\n"
        "**2. 화면 구성**\n"
        "- 목록(TPKM_FO_5_2_1) · 작성(TPKM_FO_5_2_2) · 상세+답글+댓글(TPKM_FO_5_2_3) SPA 구조.\n"
        "- 비밀글 잠금 해제 LP(TPKM_FO_5_2_4).\n\n"
        "**3. 용도**\n"
        "- 응시료 환불 신청, 회원 본인 정보(이름/생년월일/국적 등) 정정 신청.\n"
        "- 관리자 답글·댓글로 처리 상태 갱신, 작성자/관리자만 비밀글 열람.\n"
        "- 댓글/대댓글 기능 지원(고객사 0526): 비밀글 게시물의 댓글·대댓글은 자동 비밀글 처리.\n\n"
        "**4. 연동**\n"
        "- BO 환불·정보정정신청 관리(TPKM_BO_4_3) 패널.\n\n"
        "**5. 비고**\n"
        "- 글 삭제·신고 등 부가 기능은 정책 합의 후."
    ),

    # ── 환불·정보정정신청 상세 + 답글 ────────────────────────────────────
    "TPKM_FO_5_2_3_0_0_S": (
        "**1. 섹션 개요**\n"
        "- 환불·정보정정신청 상세 + 답글 + 댓글/대댓글.\n\n"
        "**2. 구성**\n"
        "- 본문 + 첨부 + 작성자 정보(마스킹).\n"
        "- 관리자 답글(공식 답변, 처리 결과 배지).\n"
        "- 댓글/대댓글 기능(고객사 0526):\n"
        "  ㄴ 작성자·관리자 모두 댓글 및 대댓글 작성 가능.\n"
        "  ㄴ 비밀글 게시물의 댓글·대댓글은 자동으로 비밀글 처리(작성자·관리자만 열람 가능).\n"
        "  ㄴ 일반 게시물의 댓글은 공개, 비공개 옵션 선택 가능.\n"
        "- 본인 글은 수정/삭제 가능(답변 완료 전).\n\n"
        "**3. 동작**\n"
        "- 비밀글은 작성자/관리자만 본문 열람.\n"
        "- 처리 상태 변경 또는 댓글/대댓글 작성 시 상대방에게 이메일 알림(문자 제외, 고객사 0526).\n\n"
        "**4. 연동**\n"
        "- BO 환불·정보정정신청 관리(TPKM_BO_4_3) 답변 폼."
    ),

    # ── 문의게시판 페이지 개요 ────────────────────────────────────────────
    "TPKM_FO_5_3_0_0_0_P": (
        "**1. 페이지 개요**\n"
        "- '게시판 · 문의게시판'(`qna.html`). 고객사 수정(0519) 신규 — 일반/비밀 글 구분, 로그인 필수.\n\n"
        "**2. 화면 구성**\n"
        "- 목록(TPKM_FO_5_3_1) · 작성(TPKM_FO_5_3_2) · 상세+답글+댓글(TPKM_FO_5_3_3) SPA.\n"
        "- 비밀글 잠금 해제 LP(TPKM_FO_5_3_4).\n\n"
        "**3. 용도**\n"
        "- 회원 → 운영진 문의(접수·시험·기타 카테고리).\n"
        "- 일반 글은 누구나 열람, 비밀 글은 작성자/관리자만.\n"
        "- 댓글/대댓글 기능 지원(고객사 0526): 비밀글 게시물의 댓글·대댓글은 자동 비밀글 처리.\n\n"
        "**4. 연동**\n"
        "- BO 문의게시판 관리(TPKM_BO_4_4)."
    ),

    # ── 문의게시판 상세 + 답글 ────────────────────────────────────────────
    "TPKM_FO_5_3_3_0_0_S": (
        "**1. 섹션 개요**\n"
        "- 문의 상세 + 답글 + 댓글/대댓글.\n\n"
        "**2. 구성**\n"
        "- 본문·첨부 + 관리자 답글(공식 답변 배지) + 상태 칩.\n"
        "- 댓글/대댓글 기능(고객사 0526):\n"
        "  ㄴ 작성자·관리자 모두 댓글 및 대댓글 작성 가능.\n"
        "  ㄴ 비밀글 게시물의 댓글·대댓글은 자동으로 비밀글 처리(작성자·관리자만 열람 가능).\n"
        "  ㄴ 일반 게시물의 댓글은 공개, 비공개 옵션 선택 가능.\n"
        "- 답변 완료 시 상태 칩 변경.\n\n"
        "**3. 동작**\n"
        "- 비밀글은 본인/관리자만 본문 열람.\n"
        "- 댓글/대댓글 작성 시 상대방에게 이메일 알림(문자 제외, 고객사 0526).\n\n"
        "**4. 연동**\n"
        "- BO 문의게시판 관리(TPKM_BO_4_4)."
    ),
}


def patch_xlsx(path: Path, page_ids: list[str]) -> int:
    if not path.is_file():
        print(f"  [SKIP] 파일 없음: {path.name}")
        return 0
    wb = load_workbook(path)
    ws = wb["기능정의"]
    updated = 0
    for row in ws.iter_rows(min_row=3):
        pid_cell  = row[7]   # 컬럼 8 = 화면 ID
        desc_cell = row[11]  # 컬럼 12 = 기능 설명
        if pid_cell.value and str(pid_cell.value).strip() in page_ids:
            pid = str(pid_cell.value).strip()
            new_desc = PATCHES.get(pid)
            if new_desc is not None:
                desc_cell.value = new_desc
                desc_cell.alignment = WRAP_TOP
                line_count = max(1, new_desc.count("\n") + 1)
                ws.row_dimensions[desc_cell.row].height = min(900, max(72, line_count * 14))
                updated += 1
                print(f"    ✔ {pid}")
    wb.save(path)
    return updated


def patch_combined(combined_path: Path, source_path: Path) -> None:
    if not combined_path.is_file():
        print(f"  [SKIP] 통합본 없음: {combined_path.name}")
        return
    all_pids = set(PATCHES.keys())
    wb = load_workbook(combined_path)
    stem = source_path.stem
    sheet_name = stem[: 31 - len("_표")] + "_표"
    if sheet_name not in wb.sheetnames:
        matches = [s for s in wb.sheetnames if s.startswith(stem[:10]) and s.endswith("_표")]
        if not matches:
            print(f"  [SKIP] 통합본에서 시트 '{sheet_name}' 찾지 못함")
            return
        sheet_name = matches[0]

    src_wb = load_workbook(source_path, data_only=False)
    src_ws = src_wb["기능정의"]
    dst_ws = wb[sheet_name]

    for src_row in src_ws.iter_rows(min_row=3):
        pid_cell  = src_row[7]
        desc_cell = src_row[11]
        if pid_cell.value and str(pid_cell.value).strip() in all_pids:
            pid = str(pid_cell.value).strip()
            for dst_row in dst_ws.iter_rows(min_row=3):
                if dst_row[7].value and str(dst_row[7].value).strip() == pid:
                    dst_row[11].value = desc_cell.value
                    dst_row[11].alignment = WRAP_TOP
                    line_count = max(1, str(desc_cell.value or "").count("\n") + 1)
                    dst_ws.row_dimensions[dst_row[11].row].height = min(900, max(72, line_count * 14))
                    print(f"    ✔ [통합본] {pid}")
                    break
    src_wb.close()
    wb.save(combined_path)


def main() -> None:
    print("=" * 60)
    print("FO 게시판 댓글/대댓글 패치 (고객사 0526)")
    print("=" * 60)

    target_file = FO_DIR / "05_게시판_기능정의서.xlsx"
    page_ids = list(PATCHES.keys())

    print(f"\n[{target_file.name}]")
    n = patch_xlsx(target_file, page_ids)
    print(f"  → {n}행 수정 완료")

    print("\n[통합본 동기화]")
    for combined_name in ["_FO_기능정의서_통합_v1.0.xlsx", "_FO_기능정의서_통합_v1.1.xlsx", "_FO_기능정의서_통합_v1.2.xlsx"]:
        combined_path = FO_DIR / combined_name
        if combined_path.is_file():
            print(f"  {combined_name}")
            patch_combined(combined_path, target_file)

    print("\n[완료]")


if __name__ == "__main__":
    main()
