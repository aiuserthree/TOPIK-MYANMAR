#!/usr/bin/env python3
"""Export 정책_합의_워크시트.md (고객사 회신용) to Excel."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "정책_합의_워크시트.xlsx"

HEADERS = ["번호", "섹션", "항목", "설명", "고객사 작성란"]

# (섹션, 항목, 설명)
ITEMS = [
    ("1. 도메인·웹사이트", "도메인 구매·등록", "topik-myanmar.com 등록 담당 업체 또는 부서"),
    ("1. 도메인·웹사이트", "DNS 설정 담당", "웹 주소 연결 설정 담당자 (성명·이메일·전화)"),
    ("1. 도메인·웹사이트", "서비스 오픈 목표일", "접수 개시일 또는 공식 공개일"),
    ("2. 이메일 발송", "발신 이메일 주소", "시스템 알림 발송 주소 (예: noreply@topik-myanmar.com) — 도메인 구매 후 확정"),
    ("2. 이메일 발송", "발신자 표시명·문의 회신", "표시명 (예: TOPIK Myanmar) / 수험생 답장 수신 주소 (예: helpdesk@…)"),
    ("3. Google 간편 로그인", "사용 여부·준비 담당", "☐ 사용 ☐ 미사용 / 사용 시 Google 앱 등록 담당자 (성명·연락처)"),
    ("4. 시험 일정·운영 데이터", "회차명", "예: 제107회"),
    ("4. 시험 일정·운영 데이터", "접수 기간", "시작일 ~ 마감일 (미얀마 현지 시각)"),
    ("4. 시험 일정·운영 데이터", "응시료 수납(오프라인)", "납부 가능 기간"),
    ("4. 시험 일정·운영 데이터", "시험일", ""),
    ("4. 시험 일정·운영 데이터", "합격자 발표일", "미정이면 「미정」으로 기재"),
    ("4. 시험 일정·운영 데이터", "수험번호 공개 일시", "수험생 마이페이지에 번호가 보이기 시작하는 날·시각"),
    ("5. 응시료·수납·시험장", "응시료 (Ⅰ / Ⅱ)", "미얀마 키앗(MMK) — Ⅰ:　　　 / Ⅱ:"),
    ("5. 응시료·수납·시험장", "수납처 안내", "납부 장소·계좌·운영시간·연락처 최종 문구"),
    ("5. 응시료·수납·시험장", "시험장·지역 목록", "1차 회차 시험장 (지역·시험장명·정원 등) — 별지 첨부 가능"),
    ("6. 운영 담당 연락처", "대표 연락처", "이메일 / 전화 (업무 시간)"),
    ("6. 운영 담당 연락처", "긴급 연락처", "접수·시험 기간 비상 연락 (성명·전화)"),
    ("7. 약관·개인정보·공지", "약관·개인정보 최종 문구", "법무 검토 완료본 제공 여부 및 예정일"),
    ("7. 약관·개인정보·공지", "마케팅·공지 이메일", "접수 안내·행사 공지 등 이메일 수신 동의 대상 발송 — ☐ 사용 ☐ 미사용"),
]

# (섹션, 항목) → 고객사 작성란 사전입력 (확정된 107회 일정)
CUSTOMER_PREFILL = {
    ("4. 시험 일정·운영 데이터", "회차명"): "제107회",
    ("4. 시험 일정·운영 데이터", "접수 기간"): "2026년 7월 17일(금) ~ 21일(화)",
    ("4. 시험 일정·운영 데이터", "응시료 수납(오프라인)"): "2026년 7월 24일(금) ~ 26일(일)",
    ("4. 시험 일정·운영 데이터", "시험일"): "2026년 10월 18일(일)",
    ("4. 시험 일정·운영 데이터", "합격자 발표일"): "미정 (추후 공지)",
}

REPLY_ROWS = [
    ("회신 이메일", ""),
    ("회신 담당자 (성명·직함)", ""),
    ("회신 기한", ""),
]

APPROVAL_ROWS = [
    ("고객사 담당자", "", ""),
    ("고객사 책임자", "", ""),
]

CONFIRMED_URLS = [
    ("수험생 사이트", "https://www.topik-myanmar.com", "도메인 확정·미구매"),
    ("관리자 페이지", "https://admin.topik-myanmar.com", "도메인 확정·미구매"),
]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize(ws, widths=None, default_max=50):
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        width = 12
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                width = max(width, min(default_max, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width


def wrap_sheet(ws, start_row=2):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)


def main():
    wb = Workbook()

    ws_guide = wb.active
    ws_guide.title = "안내"
    guide = [
        ["운영 준비 확인서 (TOPIK Myanmar)"],
        [],
        ["아래 항목을 작성·체크해 회신해 주시면 운영 준비를 진행합니다."],
        [],
        ["대상: 주미얀마 대사관 (TOPIK Myanmar 운영 담당)"],
        ["작성일: 2026-06-07"],
        [],
        ["• 화면·접수 절차는 별도 기능정의서에 협의된 내용을 따릅니다."],
        ["• 본 양식은 아직 확정이 필요한 운영 정보만 담았습니다."],
        [],
        ["확정된 서비스 주소 (참고)"],
        ["구분", "주소", "현재 상태"],
        *CONFIRMED_URLS,
        [],
        ["도메인 구매 후 웹·메일 주소 연결은 개발팀과 함께 진행합니다."],
        [],
        ["입력 항목은 「확인항목」 시트에 있습니다."],
    ]
    for line in guide:
        ws_guide.append(list(line) if isinstance(line, tuple) else line)
    autosize(ws_guide)

    ws_form = wb.create_sheet("확인항목")
    ws_form.append(HEADERS)
    for i, (section, item, desc) in enumerate(ITEMS, 1):
        ws_form.append([i, section, item, desc, CUSTOMER_PREFILL.get((section, item), "")])
    style_header(ws_form)
    wrap_sheet(ws_form)
    autosize(ws_form, widths={"A": 6, "B": 22, "C": 22, "D": 48, "E": 36})

    start = ws_form.max_row + 2
    ws_form.cell(start, 1, "회신 방법").font = Font(bold=True)
    ws_form.append(["항목", "", "", "설명", "고객사 작성란"])
    for item, desc in REPLY_ROWS:
        ws_form.append(["", "회신", item, desc, ""])
    ws_form.append([])
    appr_row = ws_form.max_row + 1
    ws_form.cell(appr_row, 1, "확인·승인").font = Font(bold=True)
    ws_form.append(["구분", "", "", "성명", "일자"])
    for role, name, date in APPROVAL_ROWS:
        ws_form.append(["", "승인", role, name, date])
    wrap_sheet(ws_form, start_row=2)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Customer input items: {len(ITEMS)}")


if __name__ == "__main__":
    main()
