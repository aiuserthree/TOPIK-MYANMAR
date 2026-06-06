#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기능정의서 FO/BO 엑셀: 일반 독자용 표현 정리
- 영어 개발·IT 용어 → 쉬운 한글 표현
- http(s) 및 .html 등 웹 파일 경로·주소 제거
- 비고: 협의/논의/합의/조율/별도 정의 등 협력이 필요한 항목만 유지
- 다국어·적용 키 나열 등 개발자 전용 블록 제거, btn-/ic-/badge- → 한글 표현
- 관련 요구사항ID: 매칭할 ID가 없으면 빈 칸으로 두며, 이 열에는 용어 치환을 적용하지 않음
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ROOT = Path(__file__).resolve().parents[1]
FO_XLSX = ROOT / "FO" / "엑셀"
BO_XLSX = ROOT / "BO" / "엑셀"


# 긴 문구부터
PHRASES: list[tuple[str, str]] = [
    (r"\bpreventDefault\b", "기본 제출 막기"),
    (r"window\.location(\.href)?", "페이지 이동"),
    (r"display\s*:\s*none", "화면에서 숨김 처리"),
    (r"role\s*=\s*[\"']alert[\"']", "긴급 알림 역할(보조기기)"),
    (r"role\s*=\s*[\"'][^\"']+[\"']", "역할(보조기기)"),
    (r"type\s*=\s*email", "이메일 입력"),
    (r"type\s*=\s*password", "비밀번호 입력"),
    (r"type\s*=\s*date", "날짜 선택"),
    (r"type\s*=\s*[^\s,)]+", "입력 형식"),
    (r"\b디바운싱\b", "연속 입력 시 마지막 입력만 처리"),
    (r"\bdebounce\b", "연속 입력 시 마지막 입력만 처리"),
    (r"\bprefill\b", "자동 채움"),
    (r"\b리다이렉트\b|\bredirect\b", "다른 페이지로 이동"),
    (r"\baccess\s*token\b|\baccessToken\b", "접근 권한 표시값"),
    (r"\btoken\b|\b토큰\b", "일시 인증값"),
    (r"\bsession\b|\b세션\b", "로그인 유지 상태"),
    (r"\bHMAC\b", "변조 여부 검사"),
    (r"\bSaaS\b", "외부 클라우드 서비스"),
    (r"SendGrid|sendgrid", "메일 발송 업체 예시"),
    (r"\bmodal\b|\bModal\b|\b디자인 모달\b|\b모달\b", "팝업 창"),
    (r"\bconfirm\b", "브라우저 확인 창"),
    (r"\balert\b", "브라우저 알림"),
    (r"\bmobile\b(?![a-z가-힣])", "모바일"),
    (r"\bheader\b(?![a-z가-힣])", "상단 헤더"),
    (r"\bfooter\b", "하단 바닥글"),
    (r"\bnav\b|\bnavigation\b", "메뉴"),
    (r"\binput\b(?![a-z가-힣])", "입력 칸"),
    (r"\bhidden\b(?![a-z가-힣])", "숨김"),
    (r"\btoggle\b", "표시 전환"),
    (r"\bvalidate\b|\bvalidation\b", "입력값 확인"),
    (r"\bsubmit\b", "제출"),
    (r"\bendpoint\b|\bAPI\b|\bapi\b", "시스템 연결 지점"),
    (r"\bJSON\b|\bjson\b", "데이터 교환 형식"),
    (r"\bCSV\b", "표 형식 데이터(CSV)"),
    (r"\bXLSX\b|\bxlsx\b", "엑셀 파일"),
    (r"\bCSS\b|\bcss\b", "화면 꾸밈"),
    (r"\bHTML\b(?![a-z\-])", "웹 화면"),
    (r"\bJavaScript\b|\bjavascript\b|\bJS\b", "자바스크립트"),
    (r"\bSVG\b", "벡터 그림"),
    (r"(?<![a-zA-Z])flex(?![a-zA-Z])", "가로·세로 정렬"),
    (r"\bgrid\b", "격자 배치"),
    (r"\bhook\b", "연결 처리"),
    (r"\bevent\b(?![a-z가-힣])", "동작"),
    (r"\bbody\b(?![a-z가-힣])", "본문 영역"),
    (r"\blabel\b", "항목 이름"),
    (r"\bplaceholder\b", "입력 안내 문구"),
    (r"\bautocomplete\b", "자동 완성"),
    (r"\bpassword\b(?![a-z가-힣])", "비밀번호"),
    (r"(?<![@.\w])\bemail\b(?![@.\w])", "이메일"),
    (r"\bstatus\b", "상태"),
    (r"\bactive\b", "활성"),
    (r"\bapproved\b", "승인됨"),
    (r"(?<!<)\bselect\b(?![a-z가-힣])", "선택"),
    (r"\boption\b", "선택지"),
    (r"\bwindow\b(?![a-z가-힣])", "창"),
    (r"\bopen\b(?![a-z가-힣])", "열기"),
    (r"\bprint\b(?![a-z가-힣])", "인쇄"),
    (r"\brow\b", "줄"),
    (r"\bpanel\b", "패널"),
    (r"\btab\b", "탭"),
    (r"\bcard\b", "카드"),
    (r"\bstep\b", "단계"),
    (r"\bbreadcrumb\b", "경로 표시"),
    (r"\blogin\b(?![a-z가-힣])", "로그인"),
    (r"\bsignup\b|\bsign-up\b", "회원가입"),
    (r"\blogout\b", "로그아웃"),
    (r"\badmin\b(?![a-z가-힣])", "관리자"),
    (r"\bdetail\b(?![a-z가-힣])", "상세"),
    (r"\bclient\b(?![a-z가-힣])|\b클라이언트\b", "이용자 기기"),
    (r"\bserver\b(?![a-z가-힣])|\b서버\b", "서버"),
    (r"\blookup\b(?![a-z가-힣])|\b조회\b", "조회"),
    (r"\bCTA\b", "안내 버튼"),
    (r"\bLP\b(?![a-zA-Z])", "별도 안내 페이지"),
    (r"\bMP\b", "팝업"),
    (r"\bFAQ\b(?![a-z가-힣])|\bfaq\b(?![a-z가-힣])", "자주 묻는 질문"),
    (r"\bearia\b|\bARIA\b", "보조기기용 표시 속성"),
    (r"\b스크린리더\b|\b스크린 리더\b", "음성 안내 프로그램"),
    (r"\b감사 로그\b", "변경 내역 기록"),
    (r"\blog\b(?![a-z가-힣])|\blogging\b", "기록"),
    (r"\bESC\b|\bEsc\b|\bescape\b", "ESC"),
    (r"\brequired\b", "필수"),
    (r"\btm_session\b", "로그인 여부 저장값"),
    (r"\bTMProfile\b", "가입자 프로필 저장"),
    (r"\bDivider\b|\bdivider\b", "구분선"),
    (r"\btoast\b|\b토스트\b", "잠깐 표시되는 안내"),
    # STEP·자동채움·영문 슬래그 페이지명
    (r"\bprefillBanner\b|\bprefill\s*banner\b", "자동 채움 안내 띠"),
    (r"\bprefetchbanner\b|\bPrefetchBanner\b|\bprefetch\s*banner\b", "자동 채움 안내 띠"),
    (r"\bprefill\b", "자동 채움"),
    (r"\bSTEP\s*(\d+)\b", r"\1단계"),
    (r"\bSTEP\b", "단계"),
    (r"(\d+)Depth\b", r"\1단계"),
    (r"\bphotos\b|\bphoto\b(?![a-z])", "사진"),
    (r"\bmypage\b|\bMyPage\b", "내 정보 페이지"),
    (r"\btopbar\b|\bTopBar\b", "맨 위 바"),
    (r"\bdashboard\b|\bDashboard\b", "현황 요약 화면"),
    (r"\bnotices\b|\bnotice\b(?![a-z가-힣])", "공지"),
    (r"\bresults\b", "결과"),
    (r"\brounds\b|\bround\b(?![a-z가-힣])", "회차"),
    (r"\bregister\b(?![a-z가-힣])", "접수"),
    (r"\bselected\b", "선택된"),
    (r"\bvisible\b", "보이는"),
    (r"\bcancelled\b|\bcanceled\b", "취소된"),
    (r"\bidle\b", "대기 상태"),
    (r"\bnext\b(?![a-z가-힣])", "다음"),
    (r"\bprev\b|\bprevious\b", "이전"),
    (r"\bshow\b(?![a-z가-힣])", "보이기"),
    (r"\bclose\b(?![a-z가-힣])", "닫기"),
    (r"\bload\b(?![a-z가-힣])", "불러오기"),
    (r"\bclick\b", "누르기"),
    (r"\bdisplay\b(?![a-z가-힣])", "표시"),
    (r"\bdate\b(?![a-z가-힣])", "날짜"),
    (r"\bauto\b(?![a-z가-힣])", "자동"),
    (r"\bfont\b(?![a-z가-힣])", "글꼴"),
    (r"\benv\b(?![a-z가-힣])", "설정 환경"),
    (r"\bmedia\b(?![a-z가-힣])", "미디어"),
    (r"\bintro\b", "소개"),
    (r"\bpage\b(?![a-z가-힣])", "페이지"),
    (r"\bterms\b(?![a-z가-힣])", "약관"),
    (r"\bphone\b(?![a-z가-힣])", "전화"),
    (r"\bprofile\b(?![a-z가-힣])", "프로필"),
    (r"\bpolite\b", "정중한 표현"),
    (r"\blive\b(?![a-z가-힣])", "실시간"),
    (r"\bimage\b(?![a-z가-힣])", "그림"),
    (r"\bmax\b(?![a-z가-힣])", "최대"),
    (r"\bbottom\b(?![a-z가-힣])", "아래쪽"),
    (r"\bgold\b(?![a-z가-힣])", "금색"),
    (r"\bthis\b", "이"),
    (r"\bchart\b(?![a-z가-힣])", "차트"),
    (r"\bspace\b(?![a-z가-힣])", "공백"),
    (r"\bsecure\b(?![a-z가-힣])", "보안"),
    (r"\bEnter\b", "Enter 키"),
    (r"\bPOST\b(?![a-z])", "정보를 서버에 보내는 방식"),
    (r"\bHTTPS\b", "보안 웹 주소"),
    (r"\bECharts\b|\becharts\b", "차트 그리기 기능(예시)"),
    (r"\bwrote\b|\bwritten\b|\bwrites?\b(?![a-z])", "기록함"),
    (r"\bingress\b|\bIngress\b", "들어오는 접속"),
    (r"\bexport\b(?![a-z])", "내보냄"),
    (r"\bimport\b(?![a-z])", "가져옴"),
    (r"\bparse\b|\bparsing\b", "항목별로 나눔 해석"),
    (r"\bserialize\b|\bserialization\b", "전달용 문자열로 변환"),
    (r"\bdeserialize\b|\bDeserialization\b", "문자열을 다시 항목으로 풂"),
    (r"\ballowlist\b|\bwhitelist\b", "허용 목록"),
    (r"\bblocklist\b|\bblacklist\b", "차단 목록"),
    (r"goRegStep\b|goRoundStep\b", "다음 단계로 이동"),
    (r"fillConfirm\b|\bfill_confirm\b", "확정 요약 채우기"),
    (r"submitReg\b", "접수 제출"),
    (r"\bisSessionValid\b", "로그인 유효 여부 확인"),
    (r"\bbatchAssignExamNumbers\b", "수험 번호 한꺼번에 매기기"),
    (r"\bcloseAdminSidebar\b", "관리자 쪽 메뉴 닫기"),
    (r"\bdataURL\b", "그림을 글자로 넣는 형식"),
    (r"\bbtn\b(?![a-z가-힣-])", "버튼"),
    (r"\bJPEG\b|\bjpeg\b|\bJPG\b|\bjpg\b", "JPEG 그림"),
    (r"\bcat\b(?![a-z가-힣])", "구분"),
    (r"\bdata\b(?![a-z가-힣])", "데이터"),
    (r"\btopik\b(?![a-z])", "TOPIK"),
    (r"\bMyanmar\b", "미얀마"),
    (r"\bKorea\b", "한국"),
    (r"\bNIIED\b", "NIIED(한국 시험 기관)"),
    (r"\bTPKM\b", "TPKM"),
    (r"\b라우팅\b", "화면 연결"),
    (r"회원가입/로그인", "회원가입·로그인"),
    (r"로그인/회원가입", "로그인·회원가입"),
    (r"저장\.load\b|저장\s*\.\s*load\b", "저장 내용 불러오기"),
    (
        r"다음\s*파라미터(?:로|을|를|에)?|다음\s*parameter\b|\bnext\s*parameter\b",
        "원래 페이지 안내값으로",
    ),
    (r"노출\b", "표시"),
    # ── 추가: 일반인이 읽기 쉬운 말(긴 문자열 먼저)
    (r"Puppeteer|puppeteer", "서버 쪽 자동 화면(인쇄·PDF 예시 도구)"),
    (r"jsPDF|\bjspdf\b", "브라우저 PDF 만들기 기능(예시)"),
    (r"D-Day\b|D‑Day\b", "마감까지 남은 날·시간 표시"),
    (r"noopener\s+noreferrer|noreferrer\s+noopener", "새 창 안전 접속 옵션"),
    (r"\bnoreferrer\b|\bnoopener\b", ""),
    (r"'lang:change'|lang:change|lang-change", "언어 바뀜 알림"),
    (r"`lang:change`", "언어 바뀜 알림"),
    (r"\bRBAC\b|\brbac\b", "직무별 접근 허용"),
    (r"\bhmac\b(?![a-z])", "변조 확인용 암호"),
    (r"\bsanitize\b|\bsanitized\b|\bsanitization\b", "위험한 내용 걷어냄 처리"),
    (r"\bXSS\b|\bxss\b", "나쁜 스크립트 끼워넣기 공격"),
    (r"\bPII\b|\bpii\b", "민감 개인 정보"),
    (r"\bEXIF\b|\bexif\b", "사진에 붙은 촬영·기기 정보"),
    (r"\bTTL\b|\bttl\b", "유효 시간"),
    (r"\bcron\b|\bCRON\b", "정해진 시간에 자동 실행"),
    (r"\bsocket\.io\b|Socket\.IO\b", "실시간 통신 기능(예시)"),
    (r"\bsocket\b(?![a-z])", "실시간 양방향 연결"),
    (r"\bwebhook\b", "외부 서비스 신호 접수 주소"),
    (r"\biframe\b|\bIframe\b", "다른 페이지를 넣어 보여 주는 창"),
    (r"\bviewport\b|\bViewport\b", "화면에서 보이는 영역 기준"),
    (r"\btimezone\b|\bUTC\b(?![a-z])", "표준 시각대"),
    (r"\btimestamp\b", "찍힌 시각"),
    (r"\bsticky\b", "스크롤해도 붙어 있는"),
    (r"\boverflow\b", "넘침"),
    (r"\btruncate\b|\btruncation\b", "길면 말줄임"),
    (r"\bscroll\b|\bscrolling\b", "스크롤"),
    (r"\bhash\b(?![a-z])", "검증용 요약값"),
    (r"\bcache\b|\bcached\b|\bcaching\b", "임시 저장"),
    (r"\bmiddleware\b", "중간에서 거르는 처리"),
    (r"\brouting\b|\brouter\b", "어느 화면으로 보낼지 연결"),
    (r"\blayout\b|\bLayout\b", "화면 틀"),
    (r"\bwidget\b|\bWidget\b", "화면 조각 부품"),
    (r"\bstack\b|\bfont\s*stack\b", "글꼴 묶음"),
    (r"Noto\s+Sans\s+KR", "한글용 설계 글꼴(Noto)"),
    (r"Noto\s+Sans\s+Myanmar", "미얀마어용 설계 글꼴(Noto)"),
    (r"\bsystem-ui\b|\bsans-serif\b", "기기 기본 글꼴"),
    (r"\bfont-family\b", "글꼴 지정"),
    (r"\bmargin\b", "바깥 여백"),
    (r"\bpadding\b", "안쪽 여백"),
    (r"\bwidth\b", "너비"),
    (r"\bheight\b", "높이"),
    (r"\bmax-width\b|\bmaxwidth\b", "최대 너비"),
    (r"\bmin-width\b|\bminwidth\b", "최소 너비"),
    (r"\bmax-height\b", "최대 높이"),
    (r"\bmin-height\b", "최소 높이"),
    (r"\boutline\b", "포커스 테두리"),
    (r"\bobject-fit\b", "그림 맞춤 방식"),
    (r"\bthumbnail\b", "작은 미리보기 그림"),
    (r"\bpreview\b", "미리보기"),
    (r"\bposition\s*:\s*fixed\b", "화면에 고정"),
    (r"\bposition\s*:\s*sticky\b", "스크롤 시 고정"),
    (r"\bposition\s*:\s*absolute\b", "겹쳐 배치"),
    (r"\bposition\s*:\s*relative\b", "기준을 두고 배치"),
    (r"\bmedia\s*query\b|\b@media\b", "화면 너비·종류에 따른 표시 규칙"),
    (r"\b@media\s*print\b", "인쇄할 때"),
    (r"\bthead\b|\btbody\b|\btfoot\b", "표 머리·본문·아랫줄"),
    (r"\bnamespace\b", "이름 앞머리(구역)"),
    (r"\bmetadata\b", "부가 설명 정보"),
    (r"\bdataset\b|\bDataset\b", "화면 속성 표기"),
    (r"\bbanner\b", "맨 위 알림 띠"),
    (r"\bcallout\b|\bCallout\b", "눈에 띄는 안내 상자"),
    (r"\bpill\b", "알약 모양 뱃지"),
    (r"\bbadge\b(?![a-z])", "짧은 띠 표시"),
    (r"\bslug\b", "웹 주소용 짧은 이름"),
    (r"\bmask\b|\bmasked\b", "가림 처리"),
    (r"\bencrypt\b|\bencryption\b|\bencrypted\b", "암호화"),
    (r"\bdecrypt\b", "복호화"),
    (r"\banonymize\b|\banonymized\b", "익명 처리"),
    (r"\brevoke\b|\brevocation\b", "없앰·취소"),
    (r"\bquota\b", "허용 한도"),
    (r"\brate\s*limit\b|\bratelimit\b", "짧은 시간 내 반복 제한"),
    (r"\bthrottle\b|\bthrottling\b", "과도한 요청 자동 제한"),
    (r"\bpolling\b", "일정 간격으로 다시 물어보기"),
    (r"\bheartbeat\b", "주기적 연결 확인"),
    (r"\bkeep-alive\b|\bkeepalive\b", "연결 유지"),
    (r"\bHTTPOnly\b|\bhttponly\b", "스크립트에서 못 읽는 쿠키 옵션"),
    (r"\bSameSite\b|\bsamesite\b", "다른 사이트에서 쿠키 보내기 제한"),
    (r"\bframe\b(?![a-z])", "틀·창"),
    (r"\bresize\b|\bresizing\b", "크기 조절"),
    (r"\bcrop\b|\bcropping\b", "잘라 내기"),
    (r"\bblur\b(?![a-z])", "흐리게"),
    (r"\bfocus\b(?![a-z])", "초점·선택"),
    (r"\bhover\b", "마우스를 올렸을 때"),
    (r"\bvisited\b", "이미 방문한 링크"),
    (r"\bskip\s*to\s*content\b", "본문으로 건너뛰기"),
    (r"\blandmark\b", "화면 구역 표시"),
    (r"\bheading\b(?![a-z])", "제목 줄"),
    (r"\bSMS\b|\bsms\b", "문자 메시지"),
    (r"\bURL\b|\burl\b(?![a-z])", "웹 주소"),
    (r"\bURI\b|\buri\b(?![a-z])", "자원 식별 주소"),
    (r"\bUUID\b|\buuid\b", "고유 번호"),
    (r"\bGUID\b|\bguid\b", "고유 식별 번호"),
    (r"\bOAuth\b|\boauth\b", "다른 사이트 계정으로 로그인"),
    (r"\bSSO\b|\bsso\b", "한 번 로그인으로 여러 서비스 이용"),
    (r"\bLDAP\b|\bldap\b", "조직 계정 서버 연동"),
    (r"\bSAML\b|\bsaml\b", "기관 간 로그인 표준"),
    (r"\bIP\b(?![a-z])|\bip\s*주소\b", "인터넷 주소"),
    (r"\bDNS\b|\bdns\b", "도메인 이름 찾기"),
    (r"\bTLS\b|\bSSL\b|\bssl\b|\btls\b", "전송 암호화"),
    (r"\bWAF\b|\bwaf\b", "웹 방화벽"),
    (r"\bDDoS\b|\bddos\b", "대량 요청 공격"),
    (r"\bCSRF\b|\bcsrf\b", "위조 요청 공격"),
    (r"\bGDPR\b|\bgdpr\b", "유럽 개인정보 규정(참고)"),
    (r"\bversioning\b|\bversion\b(?![a-z])", "버전"),
    (r"\bchangelog\b", "변경 이력"),
    (r"\brelease\b(?![a-z])", "배포"),
    (r"\bdeploy\b|\bdeployment\b", "서비스 반영"),
    (r"\brollback\b", "이전 상태로 되돌리기"),
    (r"\bhotfix\b", "긴급 수정"),
    (r"\bfeature\s*flag\b|\bfeatureflag\b", "기능 켜고 끄기 스위치"),
    (r"\bA/B\b|\bab\s*test\b|\bAB\s*test\b", "두 안을 나눠 비교 시험"),
    (r"\bfallback\b", "대체 동작"),
    (r"\bgraceful\b", "무리 없이(덜 거칠게)"),
    (r"\btimeout\b|\btime-out\b", "시간 초과"),
    (r"\bretry\b|\bretries\b", "다시 시도"),
    (r"\bbatch\b", "한꺼번에 처리"),
    (r"\bqueue\b|\bqueued\b", "대기 줄"),
    (r"\bjobs?\b(?![a-z\-])", "예약·자동 처리 작업"),
    (r"\bworker\b", "작업 처리기"),
    (r"\bmigration\b|\bmigrate\b", "데이터 옮김"),
    (r"\bseed\b(?![a-z])", "초기 예시 데이터 넣기"),
    (r"\bfixture\b", "시험용 고정 데이터"),
    (r"\bmock\b|\bstub\b", "가짜 대역"),
    (r"\bstaging\b", "시험용 단계"),
    (r"\bproduction\b|\bprod\b(?![a-z])", "실제 운영"),
    (r"\bCanary\b|\bcanary\b", "소수 사용자에게 먼저 적용"),
    (r"\bblue-green\b|\bbluegreen\b", "서버를 바꿀 때 멈춤 최소화"),
    (r"\bapplicant\b|\bapplicants\b", "신청자·접수자"),
    (r"\badmit\b(?![a-z])", "수험표·등록 확인"),
    (r"\brejected\b", "반려됨"),
    (r"\bpending\b", "대기 중"),
    (r"\bpaid\b", "결제됨"),
    (r"\bexpanded\b", "펼침"),
    (r"\bcollapsed\b", "접힘"),
    (r"\blang\b(?![a-z])", "언어"),
    (r"\brel=(?![a-z])", "링크 속성"),
    (r"\bmobilenav\b|\bMobileNav\b", "모바일 메뉴"),
    (r"\bshowpanel\b|\bShowPanel\b", "패널 전환"),
    (r"\badminauth\b|\bAdminAuth\b", "관리자 인증"),
    (r"\bnameKr\b|\bname_kr\b", "한글 성명 항목"),
    (r"\bnameEn\b|\bname_en\b", "영문 성명 항목"),
    (r"\bbirthdate\b", "생년월일 항목"),
    (r"\bgender\b(?![a-z])", "성별 항목"),
    (r"\bnationality\b", "국적 항목"),
    (r"\btextContent\b|\btextcontent\b", "화면에 보이는 글자"),
    (r"\bcurrent\b(?![a-z])", "현재"),
    (r"\bdanger\b(?![a-z])", "위험 강조 표시"),
    (r"\bIOS\b|\bios\b(?![a-z가-힣])", "애플 기기 OS"),
    (r"\babort\b(?![a-z])", "중단"),
    (r"\bqrcode\b|\bQR\s*코드\b", "QR코드"),
    (r"\bPNG\b(?![a-z])", "PNG 그림 형식"),
    (r"\bwebp\b|\bWEBP\b", "웹용 그림 형식"),
]

RE_HTTP = re.compile(r"https?://[^\s\)\]>`'\"<>]+", re.I)
RE_BACKTICK_ANY = re.compile(r"`([^`]*)`")
RE_HTML_REF = re.compile(
    r"[`'\"]?(?:\\.?\.?/)?[\w\-./]+\.html(?:\?[^\s`'\"]*)?[`'\"]?", re.I
)
RE_PATH_INLINE = re.compile(
    r"(?:`/)[^`\s]+\.html`?|/`[^`/]+\.html`|`[^`]+\.(?:html|htm)`|`web/[^`\s]+`", re.I
)
RE_ANGLE_URL = re.compile(r"<[^>\s]+\.html[^>]*>")
RE_PAREN_PATH = re.compile(r"\([\s`/]*[^\s)]*\.html[^)]*\)")
REMARK_KEEP = re.compile(
    r"(협의|논의|합의\s*필요|합의\s*후|별도\s*정의|정책\s*합의|운영\s*합의"
    r"|별도\s*합의|조율|사전\s*조율|확인\s*필요|합의\s*요망)"
)


# 개발 표기가 섞인 꺾쇠 형태 태그(화면 코드) 제거
RE_ANGLE_DEV = re.compile(r"<[/a-zA-Z가-힣_][^>]{0,500}>", re.DOTALL)
RE_FO_PAGE_PAREN = re.compile(r"모든 FO 페이지\([^)]{1,260}\)")
RE_INLINE_PATH_LIST = re.compile(
    r"\([a-z][a-z0-9_/·\s\-]{8,}(?:/[a-z][a-z0-9_]*)+[a-z0-9_/·\s\-]*\)", re.I
)

# FO 라우팅 설명 속 index/guide 같은 짧은 경로 이름 → 읽기 쉬운 한글
_SLUG_ROUTE_KR: dict[str, str] = {
    "index": "메인",
    "guide": "안내",
    "rules": "규칙",
    "notice": "공지",
    "faq": "자주 묻는 질문",
    "lookup": "성적 조회",
    "login": "로그인",
    "signup": "회원가입",
    "register": "접수",
    "mypage": "내 정보",
}
RE_SLASH_SLUG_ROUTE = re.compile(
    r"(?<![a-zA-Z0-9가-힣_/])([a-z][a-z0-9_-]*(?:/[^/\n]+)+)",
    re.I,
)


def humanize_slash_route_list(s: str) -> str:
    """예: index/guide/rules/공지 → 메인·안내·규칙·공지 (영문 슬러그만 한글명으로)"""

    def repl(m: re.Match[str]) -> str:
        chunk = m.group(1).strip()
        if chunk.count("/") < 1:
            return m.group(0)
        parts = [p.strip() for p in chunk.split("/") if p.strip()]
        if len(parts) < 2:
            return m.group(0)
        if not re.match(r"[a-z]", parts[0], re.I):
            return m.group(0)
        out: list[str] = []
        for p in parts:
            lk = p.lower()
            if re.fullmatch(r"[a-z][a-z0-9_-]*", p):
                out.append(_SLUG_ROUTE_KR.get(lk, p))
            else:
                out.append(p)
        return "·".join(out)

    return RE_SLASH_SLUG_ROUTE.sub(repl, s)


# 화면 설명용: btn-/ic-/badge- 접두어를 한글 버튼·띠지·아이콘 설명으로 바꾼다.

_BTN_KR: dict[str, str] = {
    "primary": "파란 강조 버튼",
    "secondary": "보조 버튼",
    "gold": "금색 안내 버튼",
    "login": "로그인 버튼",
    "logout": "로그아웃 버튼",
    "signup": "회원가입 버튼",
    "register": "접수 버튼",
    "guide": "안내 보기 버튼",
    "back": "돌아가기 버튼",
    "print": "인쇄 버튼",
    "pdf": "PDF 저장 버튼",
    "submit": "제출 버튼",
    "cancel": "취소 버튼",
    "confirm": "확인 버튼",
    "next": "다음 버튼",
    "prev": "이전 버튼",
    "close": "닫기 버튼",
    "reset": "초기화 버튼",
    "search": "검색 버튼",
    "filter": "필터 버튼",
    "download": "받기 버튼",
    "upload": "올리기 버튼",
    "edit": "수정 버튼",
    "delete": "삭제 버튼",
    "add": "추가 버튼",
    "save": "저장 버튼",
    "export": "내보내기 버튼",
    "import": "가져오기 버튼",
    "apply": "적용 버튼",
    "detail": "상세 보기 버튼",
    "list": "목록 버튼",
}

_IC_KR: dict[str, str] = {
    "login": "로그인 그림",
    "signup": "회원가입 그림",
    "logout": "로그아웃 그림",
    "email": "이메일 그림",
    "lock": "자물쇠 그림",
    "eye": "비밀번호 보기 그림",
    "eye-off": "비밀번호 숨기기 그림",
    "alert": "경고 그림",
    "chart": "차트 그림",
    "users": "사용자 그림",
    "image": "사진 그림",
    "calendar": "달력 그림",
    "notice": "공지 그림",
    "globe": "지구본 그림",
    "register": "접수 그림",
    "guide": "안내 그림",
    "star": "별 그림",
    "trophy": "트로피 그림",
    "camera": "카메라 그림",
    "link": "링크 그림",
    "menu": "메뉴 그림",
    "check": "체크 그림",
    "info": "안내 그림",
    "warning": "주의 그림",
}

for _ko, _en in (("로그인", "login"), ("회원가입", "signup"), ("로그아웃", "logout")):
    if _en in _IC_KR and _ko not in _IC_KR:
        _IC_KR[_ko] = _IC_KR[_en]
for _ko, _en in (("로그인", "login"), ("회원가입", "signup"), ("로그아웃", "logout")):
    if _en in _BTN_KR and _ko not in _BTN_KR:
        _BTN_KR[_ko] = _BTN_KR[_en]


_BADGE_KR: dict[str, str] = {
    "new": "새 글 띠지",
    "imp": "중요 띠지",
    "normal": "일반 띠지",
    "soon": "예정 띠지",
    "open": "접수 중 띠지",
    "closed": "마감 띠지",
    "badge": "띠지",
}


def strip_paths_and_urls(s: str) -> str:
    s = RE_HTTP.sub("", s)
    s = RE_FO_PAGE_PAREN.sub("모든 이용자 화면", s)
    s = RE_ANGLE_DEV.sub("", s)
    s = RE_ANGLE_URL.sub("", s)
    s = RE_PATH_INLINE.sub("", s)
    s = RE_HTML_REF.sub("", s)

    def _bt(m: re.Match[str]) -> str:
        inner = m.group(1)
        low = inner.lower()
        if "http://" in low or "https://" in low:
            return ""
        if inner.strip().startswith(".") and "(" not in inner:
            return ""
        if re.search(r"\.html?$|/[a-z0-9_\-]+\.(html?)(?:\?|$)", inner, re.I):
            return ""
        if inner.count("/") >= 1 and len(inner) < 120:
            seg = inner.strip("/").split("/")
            if len(seg) <= 6 and any(
                "." in x for x in seg
            ) and not inner.startswith("TPKM"):
                return ""
        if inner.startswith("#") or (inner.startswith(".") and "#" not in inner):
            return ""
        if re.fullmatch(r"[a-z0-9_\-#.\[\]()]+", inner, re.I) and "=" in inner:
            return ""
        return inner.strip()

    s = RE_BACKTICK_ANY.sub(_bt, s)
    s = RE_INLINE_PATH_LIST.sub("", s)
    s = re.sub(r"\(특히\s*\)", "", s)

    lines_out: list[str] = []
    for line in s.splitlines():
        line = RE_PAREN_PATH.sub("", line)
        line = re.sub(r"[ \t]+", " ", line)
        line = line.replace("()", "").strip()
        if line.strip(" -—·"):
            lines_out.append(line)
    s = "\n".join(lines_out)
    s = re.sub(r"\n{3,}", "\n\n", s).strip()
    return s


def apply_phrases(s: str) -> str:
    for pat, repl in PHRASES:
        s = re.sub(pat, repl, s, flags=re.I)
    for old, new in (
        ("data-i18n-웹 화면", "다국어 본문 표시"),
        ("applyLanguage(lang)", "언어 적용"),
        ("applyLanguage(", "언어 적용("),
        ("toggleMenu", "모바일 메뉴 열기"),
        ("closeMenu", "모바일 메뉴 닫기"),
    ):
        s = s.replace(old, new)
    return s


def simplify_dev_residuals(s: str) -> str:
    """남아 있는 브라우저 코드·표기 속성 줄을 평술화한다."""
    s = re.sub(
        r"document\.documentElement\.dispatchEvent\(\s*new\s+CustomEvent\([^)]*\)\s*\)",
        "브라우저에 언어 변경 알림",
        s,
        flags=re.I,
    )
    s = re.sub(r"브라우저\s*저장\.", "브라우저 저장 ", s)
    s = re.sub(r"브라우저\s*저장\s+로그인", "브라우저에 저장된 로그인", s)
    s = re.sub(r"비로그인\s*시\s*로\s+다른", "비로그인 시 다른", s)
    s = re.sub(r"-\s*/\s*진입\s*가드", "- 화면 들어가기 제한", s)
    s = s.replace("안내 잠깐 표시되는 안내", "잠깐 뜨는 안내")
    s = re.sub(r"글로벌\s*내비게이션\s*바", "머리글 메뉴 줄", s)
    s = re.sub(r"\b컴포넌트\b", "구성 부분", s)
    s = re.sub(r"\ba\.logo\b", "로고 링크", s, flags=re.I)
    s = re.sub(
        r'class\s*=\s*["\'](?:활성|active)["\']',
        "현재 메뉴 표시",
        s,
        flags=re.I,
    )
    s = re.sub(r"navigator\.language", "브라우저 기본 언어", s)
    s = re.sub(r"DOMContentLoaded|ContentLoaded", "내용 로딩 직후", s)
    s = re.sub(r"\bonchange\b", "선택 변경 시", s)
    s = re.sub(r"\bwarn\b|\bWARN\b", "경고 출력", s)
    s = re.sub(r"i18next-parser|Crowdin", "외부 번역 도구 예시", s)
    s = re.sub(r"[ \t]{2,}", " ", s)
    s = re.sub(r"\[[^\]]*data[^\]]*\]", " 다국어 전환 속성 ", s)
    s = re.sub(r"font-family:\s*[^;]+;?", "", s)
    return s


def remove_developer_translation_key_sections(text: str) -> str:
    """'다국어 키 적용'·적용 키 나열처럼 개발자만 참고하면 되는 절 블록을 제거한다."""
    title_pat = (
        r"\*\*[0-9]+\.\s*[^\*\n]*(?:다국어[^\*\n]*키|키[^\*\n]*적용|적용\s*키)[^\*\n]*\*\*"
    )
    for _ in range(40):
        m = re.search(title_pat, text, re.I)
        if not m:
            break
        tail = text[m.end() :]
        m2 = re.search(r"\n\*\*[0-9]+\.", tail)
        if m2:
            text = text[: m.start()].rstrip() + tail[m2.start() :]
        else:
            text = text[: m.start()].rstrip()
    return text


def drop_residual_key_lines(s: str) -> str:
    """절 헤더는 없고 남은 키·매핑 문장 줄을 삭제한다."""
    out: list[str] = []
    for line in s.splitlines():
        t = line.strip()
        if re.search(r"적용\s*키\s*[：:]|^-\s*data[^\n]{0,40}키", t):
            continue
        if re.search(
            r"^-[^\n]*(nav_\*|hero_[a-z]|faq_[qa]|ql_\*|foot_\*)", t, re.I
        ):
            continue
        if re.search(r"^-[^\n]*~\d|/\s*faq_", t):
            continue
        out.append(line)
    return "\n".join(out)


def substitute_btn_ic_badge_tokens(s: str) -> str:
    def bt(m: re.Match[str]) -> str:
        k = re.sub(r"[^a-z0-9_가-힣-]", "", m.group(1).lower())
        latin = re.sub(r"[^a-z0-9_-]", "", m.group(1).lower())
        return _BTN_KR.get(latin, _BTN_KR.get(k, f"{m.group(1)} 버튼"))

    def ic(m: re.Match[str]) -> str:
        latin = re.sub(r"[^a-z0-9_-]", "", m.group(1).lower())
        k = re.sub(r"[^a-z0-9_가-힣-]", "", m.group(1).lower())
        return _IC_KR.get(latin, _IC_KR.get(k, "표시 그림"))

    def bg(m: re.Match[str]) -> str:
        k = re.sub(r"[^a-z0-9_-]", "", m.group(1).lower())
        return _BADGE_KR.get(k, "띠지")

    s = re.sub(r"\.?btn-([-a-z0-9가-힣_]+)", bt, s, flags=re.I)
    s = re.sub(r"\.?ic-([-a-z0-9가-힣_]+)", ic, s, flags=re.I)
    s = re.sub(r"badge-([-a-z0-9가-힣_]+)", bg, s, flags=re.I)
    s = s.replace("ic-브라우저 알림", "경고 그림")
    s = s.replace("level-btn", "급수 선택 버튼")
    s = re.sub(r"\bhamburger\b", "햄버거", s, flags=re.I)
    return s


def repair_button_placeholder_icon(s: str) -> str:
    """괄호 안에 '버튼·아이콘'만 두고 ic- 코드가 없을 때, 기능 그림 이름으로 채운다."""
    btn_v = "\ubc84\ud2bc"
    # 엑셀·원문과 동일: 마지막 음절 U+CF58 (오타 U+CF60 아님)
    icon_ph = "\uc544\uc774\ucf58"
    lab_img = {"로그인": "로그인 그림", "회원가입": "회원가입 그림", "로그아웃": "로그아웃 그림"}

    def rep_inner(m: re.Match[str]) -> str:
        lab = m.group(1).strip()
        img = lab_img.get(lab) or _IC_KR.get(
            lab.lower() if lab.isascii() else lab, "표시 그림"
        )
        return f'"{lab}"({btn_v}, {img})'

    return re.sub(
        rf'"([^"]+)"\(\s*{re.escape(btn_v)}\s*,\s*{re.escape(icon_ph)}\s*\)',
        rep_inner,
        s,
    )


def remove_dev_identifier_parentheses(s: str) -> str:
    """(hero_badge) 등 식별자만 있는 괄호 설명을 제거한다."""

    def repl(m: re.Match[str]) -> str:
        inner = m.group(1).strip()
        if re.search(r"hero_|nav_|faq_[qa]|cd_|sec_|ql_|btn-|ic-|badge-", inner):
            return ""
        if re.fullmatch(
            r"[a-z][a-z0-9_]*/?[a-z0-9_-]*(?:\s*[/,]\s*[a-z0-9_/]+)*(?:\(웹[^\)]*\))?",
            inner,
            re.I,
        ) and "_" in inner:
            return ""
        return m.group(0)

    s = re.sub(r"\(([a-z][^()]{1,280})\)", repl, s)
    return s


def sanitize_stakeholder_content(s: str) -> str:
    s = s.replace("Q화면 디자인·조작 방식CKLINKS", "퀵링크")
    s = re.sub(
        r"([가-힣a-zA-Z0-9\)\]」』])\*\*(\s*[0-9]+\.)",
        r"\1\n**\2",
        s,
    )
    s = re.sub(r"(?<![a-zA-Z가-힣])hero(?![a-zA-Z가-힣])", "히어로", s, flags=re.I)
    s = re.sub(r"CTA(?=[가-힣ㄱ-ㅎ\)])", "안내 버튼", s, flags=re.I)
    for eng, kor in (("NOTICE:", "공지:"), ("SCHEDULE:", "일정:"), ("HERO:", "히어로:")):
        s = s.replace(eng, kor)
        s = s.replace(eng.replace(":", " "), kor.replace(":", " "))
    s = s.replace("안내 버튼를", "안내 버튼을")
    s = remove_developer_translation_key_sections(s)
    s = drop_residual_key_lines(s)
    s = substitute_btn_ic_badge_tokens(s)
    s = repair_button_placeholder_icon(s)
    s = remove_dev_identifier_parentheses(s)
    s = re.sub(r"[ \t]{2,}", " ", s)
    s = re.sub(r"\n{4,}", "\n\n\n", s)
    return s.strip()


def process_text(cell: object) -> object:
    if cell is None or not isinstance(cell, str):
        return cell
    s = cell.replace("\u00a0", " ").strip()
    if not s:
        return ""
    if s.startswith("="):
        return cell  # 수식 보존
    s = strip_paths_and_urls(s)
    s = apply_phrases(s)
    s = humanize_slash_route_list(s)
    s = strip_paths_and_urls(s)
    s = simplify_dev_residuals(s)
    s = sanitize_stakeholder_content(s)
    out_lines: list[str] = []
    for line in s.splitlines():
        t = re.sub(r"[ \t]+", " ", line).strip()
        if t:
            out_lines.append(t)
    s = "\n".join(out_lines)
    s = re.sub(r"\n{3,}", "\n\n", s).strip()
    return s


def keep_remark(text: object) -> object:
    if text is None:
        return None
    if not isinstance(text, str):
        return text
    t = text.strip()
    if not t:
        return None
    return process_text(text) if REMARK_KEEP.search(t) else None


def restore_english_column_headers(ws) -> None:
    """표 컬럼명은 Role, spec 등 영어 유지(이전 양식)."""
    for row in ws.iter_rows(min_row=1, max_row=15):
        for c in row:
            if isinstance(c, MergedCell):
                continue
            v = c.value
            if v == "역할":
                c.value = "Role"
            elif v == "적용 단계":
                c.value = "spec"



def _normalize_header_cell(v: object) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v).strip())


def _header_is_related_requirements_id(v: object) -> bool:
    """'관련 요구사항ID' 헤더(헤더 셀 공백 편차 허용)."""
    return _normalize_header_cell(v).lower() == "관련요구사항id"


def process_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for name in wb.sheetnames:
        ws = wb[name]
        restore_english_column_headers(ws)
        # 비고 헤더 열 번호 탐색
        remark_col: int | None = None
        remark_header_row: int | None = None
        for r in range(1, min(ws.max_row, 25) + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if "비고" in row_vals:
                remark_col = row_vals.index("비고") + 1
                remark_header_row = r
                break
        if remark_col is None and remark_header_row is None:
            for r in range(1, min(ws.max_row, 25) + 1):
                first = ws.cell(r, 1).value
                if first not in ("역할", "Role"):
                    continue
                remark_header_row = r
                for c in range(1, ws.max_column + 1):
                    if ws.cell(r, c).value == "관련 요구사항ID":
                        remark_col = c + 1
                        break
                if remark_col:
                    h = ws.cell(r, remark_col).value
                    if h is None or str(h).strip() == "":
                        ws.cell(r, remark_col).value = "비고"
                    break

        # 관련 요구사항ID 열: 매칭할 ID가 없으면 빈 칸으로 두며 용어 치환하지 않음
        req_col: int | None = None
        req_header_row: int | None = None
        for r in range(1, min(ws.max_row, 25) + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for c in range(1, len(row_vals) + 1):
                if _header_is_related_requirements_id(row_vals[c - 1]):
                    req_col = c
                    req_header_row = r
                    break
            if req_col is not None:
                break

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                col_idx = cell.column
                if (
                    req_col
                    and req_header_row
                    and col_idx == req_col
                    and cell.row == req_header_row
                ):
                    continue
                if req_col and col_idx == req_col:
                    v = cell.value
                    if isinstance(v, str):
                        if v.strip().startswith("="):
                            continue
                        t = v.strip()
                        cell.value = t if t else None
                    continue
                if (
                    remark_col
                    and remark_header_row
                    and col_idx == remark_col
                    and cell.row == remark_header_row
                ):
                    cell.value = "비고"
                    continue
                if remark_col and col_idx == remark_col:
                    cell.value = keep_remark(cell.value)
                else:
                    cell.value = process_text(cell.value) if cell.value is not None else None
    wb.save(path)
    wb.close()


def main() -> None:
    paths = sorted(FO_XLSX.glob("*.xlsx")) + sorted(BO_XLSX.glob("*.xlsx"))
    for p in paths:
        process_workbook(p)
        print("OK", p)


if __name__ == "__main__":
    main()
