#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
아이디(이메일) 찾기 / 비밀번호 찾기 — 구글 간편가입 케이스 추가 패치.

변경 내용:
  TPKM_FO_6_1_0_0_0_P  로그인 페이지 — 아이디 찾기 링크(6_1_2) · 비밀번호 찾기 링크(6_1_3) 분리 명시
  TPKM_FO_6_1_2_0_0_LP 아이디(이메일) 찾기 — 구글 간편가입 계정 결과 케이스 + 복수 계정 처리 추가
  TPKM_FO_6_1_3_0_0_LP 비밀번호 찾기 — 상세 흐름 보강 + 구글 간편가입 계정 분기 케이스 추가

대상 파일:
  FO/06_계정_회원가입·로그인·내정보_기능정의서.xlsx
  FO/_FO_기능정의서_통합_v1.2.xlsx (존재 시)
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
FO_DIR = ROOT / "FO"

WRAP_TOP = Alignment(wrap_text=True, vertical="top")

# ──────────────────────────────────────────────────────────────────────────────
# 수정 내용 — 화면 ID : 새 기능 설명
# ──────────────────────────────────────────────────────────────────────────────
PATCHES: dict[str, str] = {

    # ── 로그인 페이지 — 아이디/비밀번호 찾기 링크 분리 명시 ──────────────
    "TPKM_FO_6_1_0_0_0_P": (
        "**1. 페이지 개요**\n"
        "- 로그인 페이지(`login.html`).\n\n"
        "**2. UI 구성**\n"
        "- 이메일 + 비밀번호 입력, 비밀번호 표시 토글, '로그인 상태 유지'.\n"
        "- 아이디(이메일) 찾기 링크(TPKM_FO_6_1_2).\n"
        "- 비밀번호 찾기 링크(TPKM_FO_6_1_3).\n"
        "- 회원가입 링크(TPKM_FO_6_2).\n"
        "- SNS 간편 로그인(고객사 0526): 구글(Google) 계정 연동 버튼.\n"
        "  ㄴ OAuth 2.0 기반 구글 로그인.\n"
        "  ㄴ 구글 계정 최초 로그인 시 기본 정보(이름·이메일) 자동 주입 후 나머지 필수 정보 추가 입력 안내.\n\n"
        "**3. 동작**\n"
        "- 로그인 성공 시 `?next=` 보존 경로로 이동, 없으면 홈.\n"
        "- 실패 시 인라인 에러(TPKM_FO_6_1_1).\n\n"
        "**4. 비고**\n"
        "- 데모: localStorage 세션. 운영 시 서버 인증.\n"
        "- SNS 간편 로그인: 구글 계정 연동(고객사 0526).\n"
        "- 구글 간편가입 사용자는 로그인 페이지 하단 구글 버튼으로 로그인(이메일+비밀번호 입력 불필요)."
    ),

    # ── 아이디(이메일) 찾기 — 구글 간편가입 케이스 추가 ─────────────────
    "TPKM_FO_6_1_2_0_0_LP": (
        "**1. 레이어 팝업 개요**\n"
        "- 로그인 페이지(`login.html`)에서 [아이디 찾기] 클릭 시 노출되는 LP.\n"
        "- 가입 시 입력한 정보와 일치하면 **아이디(이메일 주소)** 를 안내.\n\n"
        "**2. UI 구성 — 입력 폼**\n"
        "- 한글 성명 · 생년월일 · 연락처 입력 필드.\n"
        "- [아이디 확인] 버튼.\n\n"
        "**3. 결과 표시 — 일반 가입 계정**\n"
        "- 일치 시 이메일(=로그인 아이디) 마스킹 표시.\n"
        "  ㄴ 예: hong****@naver.com\n"
        "- [로그인하러 가기] 버튼 + [비밀번호 찾기] 링크 제공.\n\n"
        "**4. 결과 표시 — 구글 간편가입 계정**\n"
        "- 동일한 이름·생년월일·연락처로 조회된 계정이 구글 간편가입인 경우:\n"
        "  ㄴ 이메일 마스킹 표시 + '구글(Google) 계정으로 가입된 계정입니다.' 안내 배지 노출.\n"
        "  ㄴ [비밀번호 찾기] 링크 비표시(구글 계정은 별도 비밀번호 없음).\n"
        "  ㄴ [Google로 로그인하기] 버튼 제공.\n\n"
        "**5. 복수 계정 처리**\n"
        "- 동일 정보로 일반 가입 + 구글 가입이 모두 존재하는 경우 두 계정을 목록으로 병렬 표시.\n"
        "  ㄴ 각 계정에 가입 유형 배지(일반 / 구글) 노출.\n"
        "  ㄴ 구글 계정에는 [Google로 로그인하기] 버튼, 일반 계정에는 [로그인하러 가기] 버튼.\n\n"
        "**6. 오류 처리**\n"
        "- 미입력: '이름, 생년월일, 연락처를 모두 입력해 주세요.'\n"
        "- 미일치: '입력하신 정보와 일치하는 계정이 없습니다.'\n"
        "- 연속 실패 시 재시도 간격 제한(운영 합의).\n\n"
        "**7. 비고**\n"
        "- 운영: 서버 조회 + 이메일 마스킹 노출·조회 횟수 제한 정책 협의.\n"
        "- 구글 간편가입 계정은 비밀번호가 없으므로 비밀번호 찾기 링크 미노출."
    ),

    # ── 비밀번호 찾기 — 상세 흐름 + 구글 간편가입 분기 케이스 추가 ───────
    "TPKM_FO_6_1_3_0_0_LP": (
        "**1. 레이어 팝업 개요**\n"
        "- 로그인 페이지(`login.html`)에서 [비밀번호 찾기] 클릭 시 노출되는 LP.\n"
        "- 입력한 이메일로 가입 유형(일반/구글)을 확인 후 분기 처리.\n\n"
        "**2. UI 구성 — 이메일 입력**\n"
        "- 이메일 입력 필드 + [확인] 버튼.\n"
        "- 이메일 형식 검증.\n\n"
        "**3. 처리 흐름 — 일반 가입 계정**\n"
        "- 이메일로 가입된 일반 계정 확인 → 비밀번호 재설정 링크 이메일 발송.\n"
        "  ㄴ '입력하신 이메일로 비밀번호 재설정 링크를 발송했습니다.' 안내.\n"
        "  ㄴ 링크 유효시간: 30분(운영 합의).\n"
        "- 재설정 링크 클릭 → 새 비밀번호 입력 폼 → 저장.\n"
        "  ㄴ 새 비밀번호는 8자 이상, 대/소/숫자/특수 조합 권장.\n\n"
        "**4. 처리 흐름 — 구글 간편가입 계정**\n"
        "- 이메일로 조회 시 구글 간편가입 계정인 경우:\n"
        "  ㄴ 비밀번호 재설정 링크 미발송.\n"
        "  ㄴ '입력하신 이메일은 Google 계정으로 가입된 계정입니다. 구글 로그인을 이용해 주세요.' 안내.\n"
        "  ㄴ [Google로 로그인하기] 버튼 제공.\n\n"
        "**5. 오류 처리**\n"
        "- 미입력: '이메일을 입력해 주세요.'\n"
        "- 미일치(가입 내역 없음): '입력하신 이메일로 가입된 계정이 없습니다.'\n"
        "- 5회 연속 실패 시 일정 시간 잠금(운영 합의).\n\n"
        "**6. 비고**\n"
        "- 메일 발송 채널(SMTP, SES 등) 운영 합의 필수.\n"
        "- 구글 간편가입 계정은 Google이 비밀번호를 관리하므로 시스템에서 재설정 불가."
    ),
}

# ──────────────────────────────────────────────────────────────────────────────
FILE_PATCHES: list[tuple[Path, list[str]]] = [
    (FO_DIR / "06_계정_회원가입·로그인·내정보_기능정의서.xlsx", list(PATCHES.keys())),
]


def patch_xlsx(path: Path, page_ids: list[str]) -> int:
    if not path.is_file():
        print(f"  [SKIP] 파일 없음: {path.name}")
        return 0

    wb = load_workbook(path)
    ws = wb["기능정의"]
    updated = 0

    for row in ws.iter_rows(min_row=3):
        pid_cell = row[7]
        desc_cell = row[11]

        if pid_cell.value and str(pid_cell.value).strip() in page_ids:
            pid = str(pid_cell.value).strip()
            new_desc = PATCHES.get(pid)
            if new_desc is not None:
                desc_cell.value = new_desc
                desc_cell.alignment = WRAP_TOP
                line_count = max(1, new_desc.count("\n") + 1)
                ws.row_dimensions[desc_cell.row].height = min(900, max(72, line_count * 14))
                updated += 1
                print(f"    ✔ {pid}")

    wb.save(path)
    return updated


def patch_combined(combined_path: Path, source_files: list[Path]) -> None:
    if not combined_path.is_file():
        print(f"  [SKIP] 통합본 없음: {combined_path.name}")
        return

    all_pids = set(PATCHES.keys())
    wb = load_workbook(combined_path)

    for src_path in source_files:
        if not src_path.is_file():
            continue
        stem = src_path.stem
        sheet_name = stem[: 31 - len("_표")] + "_표"
        if sheet_name not in wb.sheetnames:
            matches = [s for s in wb.sheetnames if s.startswith(stem[:10]) and s.endswith("_표")]
            if not matches:
                continue
            sheet_name = matches[0]

        src_wb = load_workbook(src_path, data_only=False)
        src_ws = src_wb["기능정의"]
        dst_ws = wb[sheet_name]

        for src_row in src_ws.iter_rows(min_row=3):
            pid_cell = src_row[7]
            desc_cell = src_row[11]
            if pid_cell.value and str(pid_cell.value).strip() in all_pids:
                pid = str(pid_cell.value).strip()
                for dst_row in dst_ws.iter_rows(min_row=3):
                    if dst_row[7].value and str(dst_row[7].value).strip() == pid:
                        dst_row[11].value = desc_cell.value
                        dst_row[11].alignment = WRAP_TOP
                        line_count = max(1, str(desc_cell.value or "").count("\n") + 1)
                        dst_ws.row_dimensions[dst_row[11].row].height = min(900, max(72, line_count * 14))
                        print(f"    ✔ [통합본] {pid}")
                        break
        src_wb.close()

    wb.save(combined_path)


def main() -> None:
    print("=" * 60)
    print("아이디/비밀번호 찾기 — 구글 간편가입 케이스 패치 시작")
    print("=" * 60)

    total = 0
    for path, page_ids in FILE_PATCHES:
        print(f"\n[{path.name}]")
        n = patch_xlsx(path, page_ids)
        total += n
        print(f"  → {n}행 수정 완료")

    print(f"\n{'=' * 60}")
    print(f"개별 파일 합계: {total}행 수정")
    print("=" * 60)

    print("\n[통합본 동기화]")
    fo_source = FO_DIR / "06_계정_회원가입·로그인·내정보_기능정의서.xlsx"
    fo_combined_candidates = [
        FO_DIR / "_FO_기능정의서_통합_v1.1.xlsx",
        FO_DIR / "_FO_기능정의서_통합_v1.2.xlsx",
    ]
    for fo_combined in fo_combined_candidates:
        if fo_combined.is_file():
            print(f"\n  {fo_combined.name}")
            patch_combined(fo_combined, [fo_source])

    print("\n[완료] 패치 완료.")


if __name__ == "__main__":
    main()
