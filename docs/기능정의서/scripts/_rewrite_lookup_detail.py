"""07_수험번호조회_lookup_기능정의서.xlsx의 '상세설명' 시트를
수험번호 조회 페이지(lookup.html)에 맞게 다시 작성한다.

기존 시트 서식(헤더 채움색, 굵게, 테두리, 정렬, 줄바꿈, 컬럼 폭)은 그대로 유지하고
본문만 새 내용으로 교체한다."""
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "FO" / "07_수험번호조회_lookup_기능정의서.xlsx"

SECTIONS: list[tuple[str, list[str]]] = [
    (
        "화면 흐름도",
        [
            "[수험번호 조회 페이지 진입] (lookup.html · 비로그인 가능)",
            "↓",
            "[안내 카드 + 조회 폼]",
            "↓",
            "입력: 영문 성명(여권 표기) / 생년월일(YYYYMMDD)",
            "↓",
            "[수험번호 조회] 버튼 클릭 또는 ENTER",
            "↓",
            "클라이언트 검증 (필수 / 형식)",
            "↓",
            "실패 → 인라인 에러 메시지 + 해당 입력란 포커스",
            "성공 → 응시자(이름+생년월일) 매칭 시도",
            "↓",
            "┌──────────────┬──────────────┬──────────────┐",
            "↓                                  ↓                                  ↓",
            "[성공: 수험번호 결과 카드]   [접수 확인·번호 미부여 안내]   [불일치: 일반화된 안내]",
            "수험번호 / 회차 / 급수 / 시험장   회차·급수·시험장 + “미부여” 표시   “일치하는 응시 정보가 없습니다”",
            "↓",
            "액션: 수험번호 복사 · 다시 조회 · 수험표 보기(로그인 안내)",
        ],
    ),
    (
        "주요 검증 사항",
        [
            "### 1. 본인 인증",
            "- 영문 성명(여권 표기와 동일) + 생년월일(YYYYMMDD)이 모두 일치할 때만 결과 표시",
            "- 불일치 시 일반화된 메시지(“일치하는 응시 정보가 없습니다.”) 노출 — 회원 존재 여부 등 세분화 정보는 표시하지 않음",
            "- 시도 횟수 제한: IP/세션 기준 단시간 N회 실패 시 일시 차단(정책 합의)",
            "- 향후: OTP(SMS·이메일) 인증 추가 또는 회원 로그인 필수화 검토",
            "",
            "### 2. 입력값 검증",
            "- 영문 성명: 필수 · 영문 알파벳/공백만 허용 · 대소문자·연속 공백을 정규화한 뒤 비교",
            "- 생년월일: 필수 · 8자리 숫자(YYYYMMDD) · 실제 존재하는 날짜인지 확인",
            "- 미입력·형식 오류 시 인라인 에러 + aria-live 안내",
            "",
            "### 3. 결과 상태 분기 (정합성)",
            "- 성공(수험번호 부여 완료): 결과 카드(TPKM_FO_8_1_1) 노출",
            "- 접수 확인·수험번호 미부여: 별도 안내 영역에 회차/급수/시험장만 표시하고 수험번호는 “미부여”로 명시",
            "- 미일치/없음: 에러 알림 컴포넌트(TPKM_FO_8_1_2) 노출",
            "",
            "### 4. 개인정보 보호",
            "- 결과는 본인 수험번호만 표시, 타인 식별 정보(주소·연락처 등) 노출 금지",
            "- 결과 페이지 이탈 시 화면 캐시·자동완성 잔존을 막기 위해 폼 autocomplete 정책 적용",
            "",
            "### 5. 접근성",
            "- 폼 라벨 명시 · aria-required · 결과·에러 영역에 aria-live=\"polite\"/\"assertive\" 분리",
            "- 색상 외에도 아이콘·텍스트로 성공/미부여/실패 상태를 구분",
        ],
    ),
    (
        "연동 페이지 및 기능",
        [
            "| 연동 항목 | 위치 | 비고 |",
            "|---|---|---|",
            "| 응시자(접수) 데이터 | 조회 매칭 | BO 접수관리(TPKM_BO_REQ_002) — 이름+생년월일 키 매칭 |",
            "| 회차/시험장 데이터 | 결과 카드 | BO 시험관리(TPKM_BO_REQ_004) — 활성/최신 회차 기준 |",
            "| 수험번호 부여 상태 | 결과 분기 | BO 수험번호 일괄 부여(TPKM_BO_REQ_003) 결과를 반영 |",
            "| 마이페이지 / 수험표 | 결과 액션 | 로그인 후 admit.html로 이동(REQ_020 안내) |",
            "| 공지사항 | 안내 문구 | 수험번호 부여 일정 공지로 유도 |",
            "| FAQ Q1 | 폼 하단 안내 | “수험번호는 언제 부여되나요?”로 연결 |",
        ],
    ),
    (
        "기술적 고려사항",
        [
            "### 1. 보안",
            "- 무차별 대입 방지(Rate Limiting): IP 기준 시간당 N회 제한, 임계 초과 시 캡차 또는 일시 차단",
            "- 결과는 본인 확인 직후 1회 표시 · 페이지 이탈 시 재인증 유도",
            "- 조회 파라미터는 URL/쿼리에 노출하지 않음(POST · CSRF 토큰)",
            "- 조회 시도/성공/실패 로그 기록(이름·생년월일은 마스킹)",
            "",
            "### 2. 본인 인증 강화 옵션 (향후)",
            "- 옵션 A: 휴대전화 SMS OTP — 등록된 번호로 인증코드 발송",
            "- 옵션 B: 회원 로그인 필수화 — 비로그인 조회 차단",
            "- 옵션 C: 외부 본인인증 모듈(PASS 등) 연동",
            "",
            "### 3. 다국어",
            "- 폼 라벨/플레이스홀더/안내문/에러 메시지는 한국어·미얀마어·영어 i18n 키로 관리",
            "- 결과 상태 라벨(성공/미부여/실패) 다국어 키 분리 — 예: lookup_pending_lead, res_exam_pending",
            "",
            "### 4. 성능",
            "- 단일 행 조회 위주이므로 부하 낮음 · 응답 캐시는 사용하지 않음(개인정보)",
            "- 회차 정보 등 정적 메타는 클라이언트 캐시(localStorage 5분) 가능",
            "- 1차 데모는 클라이언트 측 임시 데이터로 동작, 운영 단계에서 서버 API로 전환",
            "",
            "### 5. 데이터 라이프사이클",
            "- 조회 가능 시점: 해당 회차의 수험번호 부여가 시작된 이후 ~ 시험일 + N일까지(정책 합의)",
            "- 보존 기간 만료 후 조회 시 안내(“회차 데이터가 종료되었습니다. 본부에 문의해 주세요.”)",
            "- 동일인이 다회차 접수 이력을 가진 경우 표시 정책(최신 1건 / 회차 선택) 합의 필요",
        ],
    ),
    (
        "참고 사항",
        [
            "### 향후 기능 제안",
            "- 결과 카드에서 수험표(admit card) PDF 다운로드/인쇄 바로가기",
            "- 수험번호 부여 시 등록 이메일·SMS 자동 알림",
            "- 다회차 접수자에 대한 회차별 수험번호 리스트 보기",
            "- 시험장 약도/오시는 길 카드를 결과 영역에 통합",
            "- 캡차 / SMS OTP 단계 도입(무차별 조회 방지)",
            "",
            "### 관련 외부 정책",
            "- TOPIK 본부 수험번호 발급·관리 정책",
            "- NIIED(국립국제교육원) 운영 가이드 — 회차/접수/수험번호 부여 일정",
            "- 미얀마 한국문화원(시행기관) 안내",
        ],
    ),
]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["상세설명"]

    # 새 행 데이터 구성: (section_label_or_None, content)
    rows: list[tuple[str | None, str]] = []
    for section, lines in SECTIONS:
        for i, line in enumerate(lines):
            rows.append((section if i == 0 else None, line))

    # 헤더 행(1행)의 셀 서식을 굵게/채움 유지 — 변경하지 않음
    # 본문은 2행부터 시작
    start = 2

    # 보존할 스타일 템플릿: 기존 시트의 2행(섹션 라벨 행)과 3행(내용 행)에서 복사
    style_label = {
        "font": copy(ws.cell(row=2, column=1).font),
        "fill": copy(ws.cell(row=2, column=1).fill),
        "alignment": copy(ws.cell(row=2, column=1).alignment),
        "border": copy(ws.cell(row=2, column=1).border),
        "number_format": ws.cell(row=2, column=1).number_format,
    }
    style_body_col_a = {
        "font": copy(ws.cell(row=3, column=1).font),
        "fill": copy(ws.cell(row=3, column=1).fill),
        "alignment": copy(ws.cell(row=3, column=1).alignment),
        "border": copy(ws.cell(row=3, column=1).border),
        "number_format": ws.cell(row=3, column=1).number_format,
    }
    style_body_col_b = {
        "font": copy(ws.cell(row=3, column=2).font),
        "fill": copy(ws.cell(row=3, column=2).fill),
        "alignment": copy(ws.cell(row=3, column=2).alignment),
        "border": copy(ws.cell(row=3, column=2).border),
        "number_format": ws.cell(row=3, column=2).number_format,
    }
    # B열 본문은 wrap_text 강제
    style_body_col_b["alignment"] = Alignment(
        horizontal=style_body_col_b["alignment"].horizontal,
        vertical="center",
        wrap_text=True,
    )

    # 기존 본문 영역(2 ~ max_row) 비우기
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, 3):
            ws.cell(row=r, column=c).value = None

    # 새 내용 채우기
    for idx, (label, content) in enumerate(rows):
        r = start + idx
        a = ws.cell(row=r, column=1)
        b = ws.cell(row=r, column=2)
        a.value = label
        b.value = content

        # 스타일 적용
        s_a = style_label if label is not None else style_body_col_a
        for cell, st in ((a, s_a), (b, style_body_col_b)):
            cell.font = copy(st["font"])
            cell.fill = copy(st["fill"])
            cell.alignment = copy(st["alignment"])
            cell.border = copy(st["border"])
            cell.number_format = st["number_format"]

        # 본문 줄에 따라 행 높이 자동(기본값) 유지
        ws.row_dimensions[r].height = None

    # 잔여 행(이전보다 짧아진 경우) 정리: 행 자체 삭제는 스타일/서식만 흔들 수 있으므로
    # 값만 비워둔다. 빈 행 높이는 18로 통일.
    last_used = start + len(rows) - 1
    for r in range(last_used + 1, max_row + 1):
        ws.row_dimensions[r].height = 18

    wb.save(XLSX)
    print(f"updated: {XLSX} (rows written: {len(rows)})")


if __name__ == "__main__":
    main()
