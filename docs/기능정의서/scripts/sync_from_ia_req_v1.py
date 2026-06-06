#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기능정의서 FO/BO xlsx를 IA v1.0(고객사 0519 반영)에 따라 재생성한다.

원칙:
- IA(메뉴구조도)에 정의된 PAGE NO.만 산출 — IA에 없는 메뉴는 기능정의서를 만들지 않는다.
- 「기능 설명」 컬럼은 detailed_descriptions에서 우선 조회, 없을 경우 IA 본문을 폴백으로 사용.
- 각 단일 기능정의서의 「기능정의」 시트 데이터를 IA 기준으로 재작성하고, 부재 시 신규 파일 생성.
- FO/BO 통합본(_FO_/_BO_)은 갱신된 단일 파일에서 INDEX + 표/상세 시트를 재구성한다.

입력:
- IA, 메뉴구조도/FO/[아이뱅크] TOPIK_Myanmar_FO_IA, 메뉴구조도_v1.0.xlsx
- IA, 메뉴구조도/BO/[아이뱅크] TOPIK_Myanmar_BO_IA, 메뉴구조도_v1.0.xlsx
- scripts/detailed_descriptions.py

출력(고객사 0519 4대 메뉴 구조 반영):
  FO/
    00_공통_GNB·푸터·언어전환·모바일메뉴_기능정의서.xlsx   ← TPKM_FO_0_*
    01_메인_홈_기능정의서.xlsx                              ← TPKM_FO_1_*
    02_TOPIK안내_기능정의서.xlsx                            ← TPKM_FO_2_*
    03_TOPIK규정_기능정의서.xlsx                            ← TPKM_FO_3_*
    04_TOPIK접수_기능정의서.xlsx                            ← TPKM_FO_4_* (접수방법/시험접수/접수확인/수험표출력)
    05_게시판_기능정의서.xlsx                               ← TPKM_FO_5_* (공지/환불·정정/문의/FAQ)
    06_계정_회원가입·로그인·내정보_기능정의서.xlsx          ← TPKM_FO_6_*
    09_외부링크_기능정의서.xlsx                             ← TPKM_FO_9_*
    _FO_기능정의서_통합_v1.0.xlsx

  BO/
    00_공통_관리자레이아웃·인증_기능정의서.xlsx             ← TPKM_BO_0_*
    01_대시보드_기능정의서.xlsx                             ← TPKM_BO_1_*
    02_접수관리_기능정의서.xlsx                             ← TPKM_BO_2_* (접수자목록 + 사진심사)
    03_시험관리_기능정의서.xlsx                             ← TPKM_BO_3_* (회차 + 시험장)
    04_콘텐츠관리_기능정의서.xlsx                           ← TPKM_BO_4_* (공지/FAQ/환불·정정/문의)
    05_회원약관관리_기능정의서.xlsx                         ← TPKM_BO_5_* (회원/약관)
    06_시스템_관리자계정·처리이력_기능정의서.xlsx           ← TPKM_BO_6_* (관리자 계정·처리 이력·사이트 보기·로그아웃)
    _BO_기능정의서_통합_v1.0.xlsx
"""

from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path
from typing import Callable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parent

sys.path.insert(0, str(Path(__file__).resolve().parent))
from detailed_descriptions import get_detailed_desc  # noqa: E402

IA_FO = REPO / "IA, 메뉴구조도" / "FO" / "[아이뱅크] TOPIK_Myanmar_FO_IA, 메뉴구조도_v1.0.xlsx"
IA_BO = REPO / "IA, 메뉴구조도" / "BO" / "[아이뱅크] TOPIK_Myanmar_BO_IA, 메뉴구조도_v1.0.xlsx"

FO_DIR = ROOT / "FO"
BO_DIR = ROOT / "BO"

HEADER_ROW = [
    "Role",
    "1단계",
    "2단계",
    "3단계",
    "4단계",
    "타입",
    "화면명",
    "화면 ID",
    "PC·MO",
    "spec",
    "접근 권한",
    "기능 설명",
    "관련 요구사항ID",
    "비고",
]

THIN = Side(style="thin", color="C5CDD9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill(start_color="003478", end_color="003478", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="003478")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")


def safe_sheet_name(label: str, used: set[str]) -> str:
    label = unicodedata.normalize("NFC", label)
    s = INVALID_SHEET.sub("_", label).replace("'", "_").strip() or "SHEET"
    s = s[:31]
    base = s
    n = 2
    while s in used:
        suf = f"_{n}"
        s = base[: 31 - len(suf)] + suf
        n += 1
    used.add(s)
    return s


def type_from_page_no(page_no: str) -> str:
    suf = page_no.rsplit("_", 1)[-1]
    return {
        "P": "페이지",
        "S": "섹션",
        "C": "컴포넌트",
        "MP": "모달",
        "LP": "레이어 팝업",
        "L": "외부 링크",
        "F": "기능",
    }.get(suf, "컴포넌트")


def fo_pc_mo(page_no: str, access: str | None, note: str | None, desc: str | None) -> str:
    blob = " ".join(str(x or "") for x in [note, desc, access, page_no])
    if "모바일" in blob or "햄버거" in blob or "768" in blob:
        return "MO"
    return "공통"


def format_desc(page_id: str, ia_desc: str | None, ia_note: str | None) -> str:
    detailed = get_detailed_desc(page_id)
    if detailed:
        return detailed.strip()
    blocks: list[str] = []
    if ia_desc:
        blocks.append("**1. 메뉴·화면 요약(IA v1.0)**\n" + str(ia_desc).strip())
    if ia_note:
        blocks.append("**2. 비고(IA v1.0)**\n" + str(ia_note).strip())
    return "\n\n".join(blocks) if blocks else ""


def load_fo_ia_rows() -> list[dict]:
    wb = load_workbook(IA_FO, read_only=True, data_only=True)
    ws = wb["FO_IA"]
    rows: list[dict] = []
    for line in ws.iter_rows(min_row=6, values_only=True):
        if not line[1] or not str(line[1]).startswith("TPKM_FO_"):
            continue
        rows.append(
            {
                "page": str(line[1]).strip(),
                "role": line[2],
                "d1": line[3],
                "d2": line[4],
                "d3": line[5],
                "d4": line[6],
                "title": line[7],
                "spec": line[8],
                "access": line[9],
                "admin": line[10],
                "desc": line[11],
                "req": line[12],
                "note": line[13] if len(line) > 13 else None,
            }
        )
    wb.close()
    return rows


def load_bo_ia_rows() -> list[dict]:
    wb = load_workbook(IA_BO, read_only=True, data_only=True)
    ws = wb["BO_IA"]
    rows: list[dict] = []
    for line in ws.iter_rows(min_row=6, values_only=True):
        if not line[1] or not str(line[1]).startswith("TPKM_BO_"):
            continue
        rows.append(
            {
                "page": str(line[1]).strip(),
                "role": line[2],
                "d1": line[3],
                "d2": line[4],
                "d3": line[5],
                "d4": line[6],
                "title": line[7],
                "spec": line[8],
                "desc": line[9],
                "req": line[10],
                "note": line[11] if len(line) > 11 else None,
            }
        )
    wb.close()
    return rows


def fo_row_to_tuple(r: dict) -> tuple:
    pc_mo = fo_pc_mo(r["page"], r["access"], r["note"], r["desc"])
    desc = format_desc(r["page"], r["desc"], r["note"])
    adm = r["admin"] or ""
    return (
        r["role"],
        r["d1"],
        r["d2"],
        r["d3"],
        r["d4"],
        type_from_page_no(r["page"]),
        r["title"],
        r["page"],
        pc_mo,
        r["spec"],
        r["access"],
        desc,
        r["req"],
        adm,
    )


def bo_row_to_tuple(r: dict) -> tuple:
    desc = format_desc(r["page"], r["desc"], r["note"])
    return (
        r["role"],
        r["d1"],
        r["d2"],
        r["d3"],
        r["d4"],
        type_from_page_no(r["page"]),
        r["title"],
        r["page"],
        "공통",
        r["spec"],
        "관리자",
        desc,
        r["req"],
        r["note"] or "",
    )


def style_spec_header(ws, title: str) -> None:
    ws.cell(1, 1, value=f"{title} — 기준: IA v1.0(고객사 0519)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER_ROW))
    ws.cell(1, 1).alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 26
    for ci, h in enumerate(HEADER_ROW, start=1):
        c = ws.cell(2, ci, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = WRAP_CENTER
        c.border = BOX
    ws.row_dimensions[2].height = 26
    widths = [9, 11, 14, 22, 22, 11, 28, 28, 8, 6, 16, 96, 26, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def write_data_rows(ws, tuples: Sequence[tuple]) -> None:
    for off, tup in enumerate(tuples, start=3):
        for ci, val in enumerate(tup, start=1):
            c = ws.cell(off, ci, value=val)
            c.alignment = WRAP_TOP
            c.border = BOX
        desc = tup[11] if len(tup) > 11 else ""
        line_count = max(1, str(desc).count("\n") + 1)
        ws.row_dimensions[off].height = min(900, max(72, line_count * 14))


def filter_fo(rows: list[dict], pred: Callable[[dict], bool]) -> list[tuple]:
    return [fo_row_to_tuple(r) for r in rows if pred(r)]


def filter_bo(rows: list[dict], pred: Callable[[dict], bool]) -> list[tuple]:
    return [bo_row_to_tuple(r) for r in rows if pred(r)]


def create_or_patch_spec(path: Path, title: str, data_tuples: Sequence[tuple]) -> None:
    if path.is_file():
        wb = load_workbook(path)
        if "기능정의" not in wb.sheetnames:
            ws = wb.create_sheet("기능정의", 0)
        else:
            ws = wb["기능정의"]
            max_r = ws.max_row
            if max_r > 2:
                ws.delete_rows(3, max_r - 2)
        if "상세설명" not in wb.sheetnames:
            wb.create_sheet("상세설명")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "기능정의"
        wb.create_sheet("상세설명")
    style_spec_header(ws, title)
    write_data_rows(ws, data_tuples)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def clone_sheet_minimal(src_ws, dst_wb, title: str) -> None:
    dst = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            dst.cell(cell.row, cell.column, value=cell.value)


def rebuild_combined(paths: list[Path], combined: Path, cover: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    ws_idx = wb.create_sheet(title=safe_sheet_name("INDEX", used))
    ws_idx.cell(1, 1, value=cover).font = TITLE_FONT
    ws_idx.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws_idx.row_dimensions[1].height = 26
    hdr = ["#", "파일", "기능정의 행수", "비고"]
    for ci, h in enumerate(hdr, start=1):
        c = ws_idx.cell(3, ci, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = WRAP_CENTER
        c.border = BOX
    r = 4
    for i, p in enumerate(paths, start=1):
        src = load_workbook(p, data_only=False)
        main = src["기능정의"]
        body = src["상세설명"] if "상세설명" in src.sheetnames else None
        n_main = max(0, main.max_row - 2)
        ws_idx.cell(r, 1, value=i).alignment = WRAP_CENTER
        ws_idx.cell(r, 2, value=p.name).alignment = WRAP_TOP
        ws_idx.cell(r, 3, value=n_main).alignment = WRAP_CENTER
        ws_idx.cell(r, 4, value="IA v1.0 (고객사 0519)").alignment = WRAP_TOP
        for ci in range(1, 5):
            ws_idx.cell(r, ci).border = BOX
        r += 1

        stem = p.stem
        mname = safe_sheet_name(stem[: 31 - len("_표")] + "_표", used)
        clone_sheet_minimal(main, wb, mname)
        if body is not None:
            bname = safe_sheet_name(stem[: 31 - len("_상세")] + "_상세", used)
            clone_sheet_minimal(body, wb, bname)
        src.close()

    ws_idx.column_dimensions["A"].width = 6
    ws_idx.column_dimensions["B"].width = 70
    ws_idx.column_dimensions["C"].width = 16
    ws_idx.column_dimensions["D"].width = 32
    wb.save(combined)


# ──────────────────────────────────────────────────────────────────────────
# 신규 구조: 파일 그룹 매핑 (고객사 0519)
# ──────────────────────────────────────────────────────────────────────────
FO_JOB_DEFS: list[tuple[str, str, Callable[[dict], bool]]] = [
    (
        "00_공통_GNB·푸터·언어전환·모바일메뉴_기능정의서.xlsx",
        "공통 (GNB · 푸터 · 언어 전환 · 모바일 · 로그인 가드)",
        lambda r: r["page"].startswith("TPKM_FO_0_"),
    ),
    (
        "01_메인_홈_기능정의서.xlsx",
        "메인(홈)",
        lambda r: r["page"].startswith("TPKM_FO_1_"),
    ),
    (
        "02_TOPIK안내_기능정의서.xlsx",
        "TOPIK 안내(시험 개요 · 시험 소개 · 문항 구성 · 평가 기준)",
        lambda r: r["page"].startswith("TPKM_FO_2_"),
    ),
    (
        "03_TOPIK규정_기능정의서.xlsx",
        "TOPIK 규정(유의 사항 · 답안 작성 요령 · 응시료 규정 · 신분증 규정)",
        lambda r: r["page"].startswith("TPKM_FO_3_"),
    ),
    (
        "04_TOPIK접수_기능정의서.xlsx",
        "TOPIK 접수(접수 방법 · 시험 접수 4단계 · 접수 확인 · 수험표 출력)",
        lambda r: r["page"].startswith("TPKM_FO_4_"),
    ),
    (
        "05_게시판_기능정의서.xlsx",
        "게시판(공지사항 · 환불·정보정정신청 · 문의게시판 · FAQ)",
        lambda r: r["page"].startswith("TPKM_FO_5_"),
    ),
    (
        "06_계정_회원가입·로그인·내정보_기능정의서.xlsx",
        "계정(로그인 · 회원가입 · 마이페이지 · 내정보수정)",
        lambda r: r["page"].startswith("TPKM_FO_6_"),
    ),
    (
        "09_외부링크_기능정의서.xlsx",
        "외부 링크(TOPIK 본부 · NIIED)",
        lambda r: r["page"].startswith("TPKM_FO_9_"),
    ),
]

BO_JOB_DEFS: list[tuple[str, str, Callable[[dict], bool]]] = [
    (
        "00_공통_관리자레이아웃·인증_기능정의서.xlsx",
        "관리자 공통(레이아웃 · 인증)",
        lambda r: r["page"].startswith("TPKM_BO_0_"),
    ),
    (
        "01_대시보드_기능정의서.xlsx",
        "대시보드",
        lambda r: r["page"].startswith("TPKM_BO_1_"),
    ),
    (
        "02_접수관리_기능정의서.xlsx",
        "접수 관리(접수자 목록 · 사진 심사 · 수험번호 13자리 부여 · 연명부 양식 엑셀 · 사진 zip 다운로드)",
        lambda r: r["page"].startswith("TPKM_BO_2_"),
    ),
    (
        "03_시험관리_기능정의서.xlsx",
        "시험 관리(회차 · 시험장)",
        lambda r: r["page"].startswith("TPKM_BO_3_"),
    ),
    (
        "04_콘텐츠관리_기능정의서.xlsx",
        "콘텐츠 관리(공지사항 · FAQ · 환불·정보정정신청 · 문의게시판)",
        lambda r: r["page"].startswith("TPKM_BO_4_"),
    ),
    (
        "05_회원약관관리_기능정의서.xlsx",
        "회원·약관 관리(회원 · 약관 · 동의 이력)",
        lambda r: r["page"].startswith("TPKM_BO_5_"),
    ),
    (
        "06_시스템_관리자계정·처리이력_기능정의서.xlsx",
        "시스템(관리자 계정 관리 · 관리자 처리 이력 · 사이트 보기 · 로그아웃)",
        lambda r: r["page"].startswith("TPKM_BO_6_"),
    ),
]


def main() -> None:
    if not IA_FO.is_file() or not IA_BO.is_file():
        raise SystemExit(f"IA 파일 없음: {IA_FO} / {IA_BO}")

    fo_ia = load_fo_ia_rows()
    bo_ia = load_bo_ia_rows()

    fo_paths: list[Path] = []
    bo_paths: list[Path] = []

    for fname, title, pred in FO_JOB_DEFS:
        p = FO_DIR / fname
        rows = filter_fo(fo_ia, pred)
        create_or_patch_spec(p, title, rows)
        fo_paths.append(p)
        print(f"[FO] {fname} → {len(rows)}행")

    for fname, title, pred in BO_JOB_DEFS:
        p = BO_DIR / fname
        rows = filter_bo(bo_ia, pred)
        create_or_patch_spec(p, title, rows)
        bo_paths.append(p)
        print(f"[BO] {fname} → {len(rows)}행")

    rebuild_combined(
        fo_paths,
        FO_DIR / "_FO_기능정의서_통합_v1.0.xlsx",
        "TOPIK Myanmar FO 기능정의서 통합본 (IA v1.0 · 고객사 0519)",
    )
    rebuild_combined(
        bo_paths,
        BO_DIR / "_BO_기능정의서_통합_v1.0.xlsx",
        "TOPIK Myanmar BO 기능정의서 통합본 (IA v1.0 · 고객사 0519)",
    )
    print("[OK] 통합본 갱신 완료.")


if __name__ == "__main__":
    main()
