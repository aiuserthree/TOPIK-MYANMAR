#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
회원가입 절차 개선 패치 스크립트.

변경 내용:
  회원가입 3단계 흐름 재구성
  - 구: STEP1 기본정보+사진 / STEP2 이메일·비번 / STEP3 약관동의
  - 신: STEP1 이메일인증(구글간편가입 포함) / STEP2 기본정보+비번+사진 / STEP3 약관동의

대상 파일:
  FO/06_계정_회원가입·로그인·내정보_기능정의서.xlsx
  FO/_FO_기능정의서_통합_v1.2.xlsx (존재 시)

동작: 화면 ID(컬럼 8) 기준으로 해당 행의 기능 설명(컬럼 12)만 교체.
      나머지 컬럼·스타일·시트는 일절 변경하지 않음.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
FO_DIR = ROOT / "FO"

WRAP_TOP = Alignment(wrap_text=True, vertical="top")

# ──────────────────────────────────────────────────────────────────────────────
# 수정 내용 — 화면 ID : 새 기능 설명
# ──────────────────────────────────────────────────────────────────────────────
PATCHES: dict[str, str] = {

    # ── 회원가입 페이지 개요 — 단계 흐름 재정의 ──────────────────────────
    "TPKM_FO_6_2_0_0_0_P": (
        "**1. 페이지 개요**\n"
        "- 회원가입(`signup.html`). 3단계 스텝 + 증명사진 등록.\n"
        "- SNS 간편가입(구글 계정 연동) 옵션 포함 — STEP1에서 선택 가능.\n"
        "- 회원등급제 없음(고객사 0526 — 일반/우수/VIP 등 등급 구분 제외).\n\n"
        "**2. 흐름**\n"
        "- (SNS 간편가입) STEP1에서 '구글로 계속하기' 클릭 → Google OAuth 인증 → 이메일 자동 인증 완료 → STEP2 기본정보(구글 이름 자동 주입) + 사진 입력 → STEP3 약관 동의 → 완료.\n"
        "- (일반 가입)\n"
        "  - STEP1 이메일 인증(TPKM_FO_6_2_1) — 이메일 입력·인증코드 확인·중복 가입 차단, 인증된 이메일 = 로그인 아이디\n"
        "  - STEP2 기본정보 + 비밀번호 + 사진(TPKM_FO_6_2_2)\n"
        "  - STEP3 약관 동의(TPKM_FO_6_2_3)\n"
        "  - 완료 모달(TPKM_FO_6_2_4)\n\n"
        "**3. 비고**\n"
        "- 인증된 이메일이 곧 회원 아이디(로그인 ID).\n"
        "- 가입 시 등록한 사진은 시험 접수 STEP3에서 그대로 사용(접수 시 추가 업로드 불가).\n"
        "- SNS 간편가입: 구글 계정 연동(고객사 0526) — STEP1에서 구글 로그인으로 이메일 자동 인증.\n"
        "- 회원등급제(VIP/우수/일반) 없음(고객사 0526)."
    ),

    # ── STEP1: 이메일 인증 + 구글 간편가입 ──────────────────────────────
    "TPKM_FO_6_2_1_0_0_S": (
        "**1. 섹션 개요**\n"
        "- STEP1: 이메일 인증 및 중복 가입 확인. 인증된 이메일이 곧 로그인 아이디(ID).\n"
        "- SNS 간편가입(구글 계정 연동)도 이 단계에서 선택 가능.\n\n"
        "**2. UI 구성 — 일반 가입**\n"
        "- 이메일 입력 필드 + [인증코드 발송] 버튼.\n"
        "  ㄴ 이메일 형식 검증.\n"
        "  ㄴ 기존 가입 이메일 입력 시 '이미 가입된 이메일입니다.' 인라인 에러(중복 가입 차단).\n"
        "- 인증코드 입력 필드(6자리) + 남은 시간 타이머(5분) + [재발송] + [확인] 버튼.\n"
        "  ㄴ 코드 일치 시 이메일 인증 완료 상태로 전환(체크 아이콘 표시).\n"
        "  ㄴ 타이머 만료 시 '인증 코드가 만료되었습니다. 재발송을 눌러주세요.' 안내.\n\n"
        "**3. UI 구성 — SNS 간편가입**\n"
        "- 구분선('또는') + [G 구글 계정으로 계속하기] 버튼.\n"
        "  ㄴ 버튼 클릭 → Google OAuth 2.0 인증 팝업.\n"
        "  ㄴ 구글 인증 성공 시 이메일 자동 인증 완료 처리 → STEP2로 자동 진행(이름·이메일 자동 주입).\n"
        "  ㄴ 이미 가입된 구글 이메일로 시도 시 '이미 가입된 계정입니다. 로그인 페이지를 이용해 주세요.' 안내.\n\n"
        "**4. 검증**\n"
        "- 이메일 인증 완료 전까지 [다음] 버튼 비활성.\n"
        "- 중복 이메일 차단(일반 가입·SNS 가입 통합 검증).\n\n"
        "**5. 비고**\n"
        "- 인증 코드는 서버에서 생성·검증(데모 환경 제외).\n"
        "- SNS 간편가입 시 구글에서 제공된 이메일이 아이디가 됨."
    ),

    # ── STEP2: 기본정보 + 비밀번호 + 사진 (기존 STEP1+STEP2 통합) ───────
    "TPKM_FO_6_2_2_0_0_S": (
        "**1. 섹션 개요**\n"
        "- STEP2: 회원 기본정보 + 비밀번호 설정 + 증명사진 등록.\n"
        "- STEP1에서 인증된 이메일은 아이디로 고정(readonly 표시, 변경 불가).\n\n"
        "**2. 입력 항목 — 기본정보**\n"
        "- 한글/영문 성명 · 생년월일 · 성별 · 국적 · 제1언어 · 연락처 · 직업(코드) · 응시동기(코드) · 응시목적(코드). (여권번호 미수집)\n"
        "- SNS 간편가입 시 구글 계정에서 제공된 이름 자동 주입(수정 가능, 이메일은 수정 불가).\n\n"
        "**3. 비밀번호 설정 (일반 가입만 해당)**\n"
        "- 비밀번호 8자 이상, 대/소/숫자/특수 조합 권장, 비밀번호 확인 일치 검증.\n"
        "- SNS 간편가입(구글) 사용자는 별도 비밀번호 설정 없음.\n"
        "- 비밀번호 변경 주기(고객사 0526): 6개월마다 변경 권고(계정 관리 및 대리접수 방지 목적). 6개월 경과 시 로그인 후 변경 안내 팝업 노출(강제 변경 또는 다음에 변경 선택 — 운영 합의).\n\n"
        "**4. 증명사진 등록**\n"
        "- jpg, 200KB~2MB 업로드, 미리보기 제공.\n"
        "- 규격 1차 검증(파일 형식/크기/비율).\n\n"
        "**5. 검증**\n"
        "- 모든 필수 항목 입력 + 사진 업로드 완료 시 [다음] 활성.\n"
        "- 직업·응시동기·응시목적 코드는 '연명부 양식' 코드와 일치(직업 1~12 / 응시동기 1~11 / 응시목적 1~15).\n\n"
        "**6. 비고**\n"
        "- 이메일 필드는 STEP1 인증 값으로 고정(수정 필요 시 이전 단계로 돌아가기)."
    ),
}

# ──────────────────────────────────────────────────────────────────────────────
# 파일별 수정 대상 page_id 매핑
# ──────────────────────────────────────────────────────────────────────────────
FILE_PATCHES: list[tuple[Path, list[str]]] = [
    (FO_DIR / "06_계정_회원가입·로그인·내정보_기능정의서.xlsx", [
        "TPKM_FO_6_2_0_0_0_P",
        "TPKM_FO_6_2_1_0_0_S",
        "TPKM_FO_6_2_2_0_0_S",
    ]),
]


def patch_xlsx(path: Path, page_ids: list[str]) -> int:
    """지정된 xlsx 파일에서 화면 ID(컬럼 8)가 일치하는 행의 기능 설명(컬럼 12)만 교체."""
    if not path.is_file():
        print(f"  [SKIP] 파일 없음: {path.name}")
        return 0

    wb = load_workbook(path)
    ws = wb["기능정의"]
    updated = 0

    for row in ws.iter_rows(min_row=3):
        pid_cell = row[7]   # 컬럼 8 (0-indexed: 7) = 화면 ID
        desc_cell = row[11]  # 컬럼 12 (0-indexed: 11) = 기능 설명

        if pid_cell.value and str(pid_cell.value).strip() in page_ids:
            pid = str(pid_cell.value).strip()
            new_desc = PATCHES.get(pid)
            if new_desc is not None:
                desc_cell.value = new_desc
                desc_cell.alignment = WRAP_TOP
                line_count = max(1, new_desc.count("\n") + 1)
                ws.row_dimensions[desc_cell.row].height = min(900, max(72, line_count * 14))
                updated += 1
                print(f"    ✔ {pid}")

    wb.save(path)
    return updated


def patch_combined(combined_path: Path, source_files: list[Path]) -> None:
    """통합 xlsx의 _표 시트를 개별 파일에서 다시 읽어 기능 설명 동기화."""
    if not combined_path.is_file():
        print(f"  [SKIP] 통합본 없음: {combined_path.name}")
        return

    all_pids = set(PATCHES.keys())
    wb = load_workbook(combined_path)

    for src_path in source_files:
        if not src_path.is_file():
            continue
        stem = src_path.stem
        sheet_name = stem[: 31 - len("_표")] + "_표"
        if sheet_name not in wb.sheetnames:
            matches = [s for s in wb.sheetnames if s.startswith(stem[:10]) and s.endswith("_표")]
            if not matches:
                continue
            sheet_name = matches[0]

        src_wb = load_workbook(src_path, data_only=False)
        src_ws = src_wb["기능정의"]
        dst_ws = wb[sheet_name]

        for src_row in src_ws.iter_rows(min_row=3):
            pid_cell = src_row[7]
            desc_cell = src_row[11]
            if pid_cell.value and str(pid_cell.value).strip() in all_pids:
                pid = str(pid_cell.value).strip()
                for dst_row in dst_ws.iter_rows(min_row=3):
                    if dst_row[7].value and str(dst_row[7].value).strip() == pid:
                        dst_row[11].value = desc_cell.value
                        dst_row[11].alignment = WRAP_TOP
                        line_count = max(1, str(desc_cell.value or "").count("\n") + 1)
                        dst_ws.row_dimensions[dst_row[11].row].height = min(900, max(72, line_count * 14))
                        print(f"    ✔ [통합본] {pid}")
                        break
        src_wb.close()

    wb.save(combined_path)


def main() -> None:
    print("=" * 60)
    print("회원가입 절차 개선 패치 시작")
    print("=" * 60)

    total = 0
    for path, page_ids in FILE_PATCHES:
        print(f"\n[{path.name}]")
        n = patch_xlsx(path, page_ids)
        total += n
        print(f"  → {n}행 수정 완료")

    print(f"\n{'=' * 60}")
    print(f"개별 파일 합계: {total}행 수정")
    print("=" * 60)

    # FO 통합본 동기
    print("\n[통합본 동기화]")
    fo_source = FO_DIR / "06_계정_회원가입·로그인·내정보_기능정의서.xlsx"
    fo_combined_candidates = [
        FO_DIR / "_FO_기능정의서_통합_v1.1.xlsx",
        FO_DIR / "_FO_기능정의서_통합_v1.2.xlsx",
    ]
    for fo_combined in fo_combined_candidates:
        if fo_combined.is_file():
            print(f"\n  {fo_combined.name}")
            patch_combined(fo_combined, [fo_source])

    print("\n[완료] 회원가입 절차 개선 패치 완료.")


if __name__ == "__main__":
    main()
