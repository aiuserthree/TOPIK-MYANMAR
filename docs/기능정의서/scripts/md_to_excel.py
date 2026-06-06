#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TOPIK Myanmar 기능정의서 Markdown → Excel 변환기

reference/기능정의서 양식을 그대로 따라:
- "기능정의" 시트: 메인 마크다운 표(Role / 1Depth ~ 비고)를 그대로 컬럼별로 분리
- "상세설명" 시트: 화면 흐름도 / 검증 / 연동 / 기술 / 참고 등 표 외 본문을 섹션·라인 단위로 출력
- 통합 워크북: 모든 md 파일을 한 워크북으로 묶고, 첫 시트에 인덱스 제공

작성 규칙:
- 메인 표의 「관련 요구사항ID」: 요구사항 정의서와 맞출 ID가 없으면 빈 칸으로 둠

출력 위치:
- FO: 기능정의서/FO/엑셀/{파일별}.xlsx + 기능정의서/FO/엑셀/_FO_기능정의서_통합.xlsx
- BO: 기능정의서/BO/엑셀/{파일별}.xlsx + 기능정의서/BO/엑셀/_BO_기능정의서_통합.xlsx
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
FO_MD = ROOT / "FO" / "md"
BO_MD = ROOT / "BO" / "md"
FO_OUT = ROOT / "FO" / "엑셀"
BO_OUT = ROOT / "BO" / "엑셀"

# Excel per-cell character limit
MAX_CELL = 32_700

# reference/기능정의서 컬럼 순서를 그대로 사용
COLS = [
    "Role",
    "1Depth",
    "2Depth",
    "3Depth",
    "4Depth",
    "타입",
    "화면명",
    "화면 ID",
    "PC/MO",
    "spec",
    "접근 권한",
    "기능 설명",
    "관련 요구사항ID",
    "비고",
]

HEADER_FILL = PatternFill(start_color="003478", end_color="003478", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="EAF1F8", end_color="EAF1F8", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="003478")
THIN = Side(border_style="thin", color="C5CDD9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

INVALID = re.compile(r"[\[\]:*?/\\]")


def split_markdown_table_row(line: str) -> list[str]:
    """
    마크다운 표 한 줄을 셀 단위로 분리한다.
    `...`, <...> 안의 파이프(|)는 열 구분자로 취급하지 않는다(기능 설명 내 HTML/코드 대응).
    """
    s = line.strip()
    if not s.startswith("|"):
        return []
    s = s[1:]
    if s.endswith("|"):
        s = s[:-1]

    cells: list[str] = []
    buf: list[str] = []
    backtick = False
    angle = 0
    i = 0
    while i < len(s):
        ch = s[i]
        if ch == "`":
            backtick = not backtick
            buf.append(ch)
            i += 1
            continue
        if not backtick:
            if ch == "<":
                # 수식/화면크기 비교용 `<7` 등은 태그로 보지 않음 (`<br`, `<div`, `</` 만 태그)
                nxt = s[i + 1] if i + 1 < len(s) else ""
                if nxt.isalpha() or nxt in "/!?":
                    angle += 1
                buf.append(ch)
                i += 1
                continue
            elif ch == ">" and angle > 0:
                angle -= 1
            elif ch == "|" and angle == 0:
                cells.append("".join(buf).strip())
                buf = []
                i += 1
                continue
        buf.append(ch)
        i += 1
    cells.append("".join(buf).strip())
    return cells


def safe_sheet_name(label: str, used: set[str]) -> str:
    # macOS HFS+/APFS는 NFD로 한글을 분해 저장 → 31자 제한 안에서 자모가 잘리지 않도록 NFC 정규화
    label = unicodedata.normalize("NFC", label)
    s = INVALID.sub("_", label).replace("'", "_").strip() or "SHEET"
    s = s[:31]
    base = s
    n = 2
    while s in used:
        suf = f"_{n}"
        s = base[: 31 - len(suf)] + suf
        n += 1
    used.add(s)
    return s


def truncate_cell(text: str) -> str:
    text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    if len(text) > MAX_CELL:
        text = text[: MAX_CELL - 3] + "..."
    return text


def parse_md(md_path: Path) -> tuple[str, list[dict[str, str]], list[tuple[str, list[str]]]]:
    """
    Returns (title, rows, body_blocks).
    - title: 첫 번째 # 헤더 텍스트
    - rows: 메인 표(첫 번째 마크다운 표)의 데이터 행 dict 리스트 (헤더 키 그대로)
    - body_blocks: 표 이후의 ## 섹션 단위 (heading, lines)
    """
    raw = md_path.read_text(encoding="utf-8")
    lines = raw.splitlines()
    n = len(lines)

    title = md_path.stem
    rows: list[dict[str, str]] = []
    body: list[tuple[str, list[str]]] = []

    i = 0
    # title
    while i < n and not lines[i].lstrip().startswith("#"):
        i += 1
    if i < n:
        title = lines[i].lstrip("#").strip()
        i += 1

    main_table_done = False

    while i < n:
        ln = lines[i].rstrip()

        if not main_table_done and ln.lstrip().startswith("| Role |"):
            header_cells = split_markdown_table_row(ln)
            i += 1
            # separator row
            if i < n and lines[i].lstrip().startswith("|") and re.search(r"-{3,}", lines[i]):
                i += 1
            # data rows
            while i < n:
                row_line = lines[i].rstrip()
                if not row_line.lstrip().startswith("|"):
                    break
                cells = split_markdown_table_row(row_line)
                while len(cells) < len(header_cells):
                    cells.append("")
                rec = {h: v for h, v in zip(header_cells, cells)}
                rows.append(rec)
                i += 1
            main_table_done = True
            continue

        if ln.startswith("## "):
            heading = ln[3:].strip()
            block: list[str] = []
            i += 1
            while i < n and not lines[i].startswith("## "):
                block.append(lines[i])
                i += 1
            body.append((heading, block))
            continue

        i += 1

    return title, rows, body


def write_main_sheet(ws, title: str, rows: list[dict[str, str]]) -> None:
    """메인 기능정의 표 (reference 양식)"""
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP_CENTER
        c.border = BOX
    ws.row_dimensions[2].height = 26

    for r_offset, rec in enumerate(rows, start=3):
        for ci, h in enumerate(COLS, start=1):
            v = rec.get(h, "")
            c = ws.cell(row=r_offset, column=ci, value=truncate_cell(v))
            c.alignment = WRAP_TOP
            c.border = BOX
        # 동적으로 행 높이를 늘림 (기능 설명 분량 기준)
        spec_text = rec.get("기능 설명", "")
        line_count = max(1, spec_text.count("<br>") + spec_text.count("\n"))
        ws.row_dimensions[r_offset].height = min(420, max(80, line_count * 14))

    widths = [9, 11, 14, 16, 16, 11, 26, 28, 8, 6, 16, 90, 24, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def write_body_sheet(ws, body: list[tuple[str, list[str]]]) -> None:
    """표 외 본문 (화면 흐름도/검증/연동/기술/참고 등)"""
    headers = ["섹션", "라인", "내용"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP_CENTER
        c.border = BOX
    ws.row_dimensions[1].height = 22

    r = 2
    for heading, block_lines in body:
        # 섹션 헤더 색깔이 들어간 한 줄
        ws.cell(row=r, column=1, value=heading).font = Font(bold=True, color="003478")
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.cell(row=r, column=2, value="").fill = SECTION_FILL
        ws.cell(row=r, column=3, value=heading).fill = SECTION_FILL
        ws.cell(row=r, column=3).font = Font(bold=True, color="003478")
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).border = BOX
            ws.cell(row=r, column=ci).alignment = WRAP_TOP
        r += 1

        for line_no, line in enumerate(block_lines, start=1):
            ws.cell(row=r, column=1, value="").alignment = WRAP_TOP
            ws.cell(row=r, column=2, value=line_no).alignment = WRAP_CENTER
            ws.cell(row=r, column=3, value=truncate_cell(line)).alignment = WRAP_TOP
            for ci in range(1, 4):
                ws.cell(row=r, column=ci).border = BOX
            r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 130
    ws.freeze_panes = "A2"


def build_workbook_for_md(md_path: Path) -> Workbook:
    title, rows, body = parse_md(md_path)
    wb = Workbook()
    used: set[str] = set()
    ws_main = wb.active
    ws_main.title = safe_sheet_name("기능정의", used)
    write_main_sheet(ws_main, title or md_path.stem, rows)

    ws_body = wb.create_sheet(title=safe_sheet_name("상세설명", used))
    write_body_sheet(ws_body, body)
    return wb


def build_combined_workbook(md_files: list[Path], cover_title: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    # INDEX 시트
    ws = wb.create_sheet(title=safe_sheet_name("INDEX", used))
    ws.cell(row=1, column=1, value=cover_title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 26

    headers = ["#", "파일", "제목", "행수"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP_CENTER
        c.border = BOX

    parsed: list[tuple[Path, str, list[dict[str, str]], list[tuple[str, list[str]]]]] = []
    for md in md_files:
        t, rows, body = parse_md(md)
        parsed.append((md, t, rows, body))

    r = 4
    for idx, (md, t, rows, _body) in enumerate(parsed, start=1):
        ws.cell(row=r, column=1, value=idx).alignment = WRAP_CENTER
        ws.cell(row=r, column=2, value=md.name).alignment = WRAP_TOP
        ws.cell(row=r, column=3, value=t).alignment = WRAP_TOP
        ws.cell(row=r, column=4, value=len(rows)).alignment = WRAP_CENTER
        for ci in range(1, 5):
            ws.cell(row=r, column=ci).border = BOX
        r += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 64
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 8

    # 각 md 파일별 시트 (메인 표 + 상세설명을 분리해서 시트 2개씩)
    # 31자 제한 안에서 메인/상세 구분이 보이도록 접미어를 stem 앞에서 잘라냄
    def _short_stem(stem: str, suffix: str) -> str:
        stem = unicodedata.normalize("NFC", stem)
        head_max = 31 - len(suffix)
        return stem[:head_max] + suffix

    for md, t, rows, body in parsed:
        main_sname = safe_sheet_name(_short_stem(md.stem, "_표"), used)
        ws_main = wb.create_sheet(title=main_sname)
        write_main_sheet(ws_main, t or md.stem, rows)

        body_sname = safe_sheet_name(_short_stem(md.stem, "_상세"), used)
        ws_body = wb.create_sheet(title=body_sname)
        write_body_sheet(ws_body, body)

    return wb


def main() -> None:
    fo_files = sorted(FO_MD.glob("*.md"))
    bo_files = sorted(BO_MD.glob("*.md"))

    FO_OUT.mkdir(parents=True, exist_ok=True)
    BO_OUT.mkdir(parents=True, exist_ok=True)

    print(f"[FO] md {len(fo_files)}건 → {FO_OUT.relative_to(ROOT)}")
    for md in fo_files:
        wb = build_workbook_for_md(md)
        out = FO_OUT / (INVALID.sub("_", md.stem) + ".xlsx")
        wb.save(out)
        print(f"  - {out.name}")

    print(f"[BO] md {len(bo_files)}건 → {BO_OUT.relative_to(ROOT)}")
    for md in bo_files:
        wb = build_workbook_for_md(md)
        out = BO_OUT / (INVALID.sub("_", md.stem) + ".xlsx")
        wb.save(out)
        print(f"  - {out.name}")

    fo_combined = build_combined_workbook(fo_files, "TOPIK Myanmar FO 기능정의서 통합본")
    fo_combined_path = FO_OUT / "_FO_기능정의서_통합.xlsx"
    fo_combined.save(fo_combined_path)
    print(f"[FO] 통합본: {fo_combined_path.name}")

    bo_combined = build_combined_workbook(bo_files, "TOPIK Myanmar BO 기능정의서 통합본")
    bo_combined_path = BO_OUT / "_BO_기능정의서_통합.xlsx"
    bo_combined.save(bo_combined_path)
    print(f"[BO] 통합본: {bo_combined_path.name}")

    print("Done.")


if __name__ == "__main__":
    main()
