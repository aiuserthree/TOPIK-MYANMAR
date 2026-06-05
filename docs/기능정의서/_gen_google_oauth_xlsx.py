#!/usr/bin/env python3
"""Generate customer-facing Excel from 고객사_확인요청_도메인_구글간편가입.md tables."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT = "/Users/jhcho/Documents/Myanmar_v2.0/docs/기능정의서/고객사_확인요청_도메인_구글간편가입.xlsx"

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP


def write_table(ws, headers, rows, widths):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, widths)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP


def main():
    wb = Workbook()

    ws = wb.active
    ws.title = "안내"
    ws["A1"] = "고객사 확인 요청 — 도메인·DNS·Google 간편가입"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "발신: 개발팀"
    ws["A4"] = "수신: 주미얀마 대사관·MOFA IT 담당"
    ws["A5"] = "작성일: 2026-06-05"
    ws["A7"] = "회신 방법"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = "각 시트의 「확정값(회신)」 열을 채워 회신해 주세요."
    ws["A9"] = "선택지가 있는 항목은 옵션 A / B 중 하나를 명시해 주세요."
    ws["A11"] = "권장 일정"
    ws["A11"].font = Font(bold=True)
    ws["A12"] = "· 오픈 D-7: 루트·서브도메인 FQDN, DNS 담당자, 오픈 목표일 확정"
    ws["A13"] = "· 오픈 D-3: DNS 레코드 등록 완료 (전파 24~48시간)"
    ws["A15"] = "백엔드·개발 환경 안내"
    ws["A15"].font = Font(bold=True)
    ws["A16"] = "API는 고객사 Python FastAPI 자체 개발·통합 운영 — 별도 api. 서브도메인 확정 항목 아님"
    ws["A17"] = "별도 dev FQDN·서버 없음 — Vercel Preview·로컬에서 개발·QA"
    ws["A19"] = "Google OAuth"
    ws["A19"].font = Font(bold=True)
    ws["A20"] = "· Google 간편가입 도입 확정 (도입 여부 재확인 불필요)"
    ws["A21"] = "· 단일 Client ID — Vercel URL 먼저 등록, prod FQDN 확정 후 동일 클라이언트에 origin 추가"
    ws["A22"] = "· 개인정보처리방침 URL 운영 필수 — 고객사(법무) 최종 URL 회신"
    ws["A23"] = "· Cloud 프로젝트: 대사관 소유 권장 + 개발팀 Editor 또는 Client ID 전달"
    ws.column_dimensions["A"].width = 80

    ws1 = wb.create_sheet("도메인_DNS")
    write_table(
        ws1,
        ["#", "항목", "결정옵션", "권장·기본", "담당", "확정값(회신)"],
        [
            ["1-1", "루트 도메인", "예: topik-myanmar.gov.mm, topik.mm, 대사관 기존 도메인 하위", "MOFA·대사관 공식 도메인 1개 확정", "고객사", ""],
            ["1-2", "도메인 등록·갱신", "대사관 자체 / MOFA 중앙 / 위탁 업체", "등록 주체·만료일·갱신 담당 명시", "고객사", ""],
            ["1-3", "DNS 관리", "Cafe24, Route53, Cloudflare, 사내 DNS 등", "IT 담당자 성명·이메일·전화", "고객사 IT", ""],
            ["1-4", "개발팀 권한", "DNS 읽기만 / 레코드 등록 위임 / 서브도메인 위임", "값 제공 후 IT가 등록, 개발팀은 검증만", "고객사 IT + 개발팀", ""],
            ["1-5", "오픈 목표일", "접수 개시일 또는 서비스 공개일", "DNS D-7 확정, D-3 등록 완료", "고객사", ""],
        ],
        {"A": 6, "B": 18, "C": 38, "D": 32, "E": 16, "F": 28},
    )

    ws2 = wb.create_sheet("FO_BO_URL")
    write_table(
        ws2,
        ["#", "서비스", "결정 필요 사항", "옵션 A (권장)", "옵션 B", "담당", "확정값(회신)"],
        [
            ["2-1", "FO (수험생)", "프로덕션 FQDN", "https://www.____________ (apex→www 리다이렉트)", "https://____________ (non-www 기준)", "고객사 + 개발", ""],
            ["2-2", "www 정책", "apex vs www", "www를 정식, apex는 301→www", "non-www 정식", "고객사", ""],
            ["2-3", "BO (관리자)", "별도 호스트 vs 동일 origin", "https://admin.____________ (별도 웹 호스팅)", "https://www.____________/admin", "고객사 + 개발", ""],
            ["2-4", "staging", "중간 환경 필요 여부", "없음", "별도 staging 도메인", "고객사", ""],
        ],
        {"A": 6, "B": 14, "C": 22, "D": 38, "E": 32, "F": 16, "G": 28},
    )
    ws2.append([])
    ws2.append(["개발·시연: 별도 dev FQDN 없음 — Vercel Preview·로컬"])
    ws2.append([])
    ws2.append(["현재 임시 URL (참고)"])
    ws2.append(["서비스", "임시 URL"])
    ws2.append(["FO", "https://topik-myanmar.vercel.app"])
    ws2.append(["BO", "https://topik-myanmar-c-bo.vercel.app"])

    ws3 = wb.create_sheet("이메일_DNS")
    write_table(
        ws3,
        ["#", "항목", "결정 필요 사항", "권장", "담당", "확정값(회신)"],
        [
            ["3-1", "발신 도메인", "no-reply@____________ 실제 도메인", "MOFA/대사관 공식 도메인", "고객사", ""],
            ["3-2", "발신명·Reply-To", "표시명, 회신 주소", "TOPIK Myanmar / helpdesk@____________", "고객사 운영", ""],
            ["3-3", "SPF·DKIM·DMARC", "DNS TXT/CNAME — 메일 서비스 대시보드 값", "오픈 전 Verified 필수; 초기 DMARC p=none 검토", "고객사 IT 등록, 개발 검증", ""],
            ["3-4", "dev 발송", "실수 발송 방지", "Mailtrap·MAIL_PROVIDER=console", "개발팀 (참고)", "개발팀 처리"],
        ],
        {"A": 6, "B": 16, "C": 32, "D": 36, "E": 22, "F": 28},
    )

    ws4 = wb.create_sheet("Google_간편가입")
    write_table(
        ws4,
        ["#", "항목", "설명", "담당", "확정값(회신)"],
        [
            ["4-1", "Google Cloud 프로젝트", "대사관 소유 권장. 개발팀 Editor 권한 또는 Console 설정 후 Client ID 전달", "고객사", ""],
            ["4-2", "동의 화면 — 앱명·로고", "OAuth 브랜딩 (미얀마 TOPIK 접수, scope: openid·email·profile)", "고객사", ""],
            ["4-3", "개인정보처리방침 URL", "운영 필수 — Google Production 검증·동의 화면. FO privacy.html과 동일", "고객사 + 법무", ""],
            ["4-4", "FO URL (§2 확정 후)", "확정 FO FQDN — 동일 OAuth 클라이언트에 origin 추가", "개발팀 등록, 고객사 승인", ""],
        ],
        {"A": 6, "B": 24, "C": 52, "D": 18, "E": 32},
    )
    ws4.append([])
    ws4.append(["요청하지 않음: Google 도입 여부, 별도 dev FQDN, dev/prod Client ID 분리"])
    ws4.append([])
    ws4.append(["단일 Client ID — 현재 임시 origins"])
    ws4.append(["서비스", "URL"])
    ws4.append(["FO", "https://topik-myanmar.vercel.app"])
    ws4.append(["BO (Google 로그인 시)", "https://topik-myanmar-c-bo.vercel.app"])

    ws5 = wb.create_sheet("역할분담")
    write_table(
        ws5,
        ["작업", "고객사 / IT", "개발팀"],
        [
            ["루트·서브도메인 FQDN 확정 (FO·BO)", "●", "○ (기술 제안)"],
            ["DNS 레코드 등록 (A/CNAME/TXT)", "●", "○ (웹 호스팅·메일 서비스에서 값 발급·요청서 작성)"],
            ["메일 발송 도메인 Verified", "● (DNS)", "○ (메일 서비스 대시보드·발신 설정)"],
            ["Google Cloud 프로젝트 소유·Editor 권한(또는 Client ID 전달)", "●", "○ (Console 연동·origin 등록)"],
            ["Google 동의 화면(앱명·로고)·개인정보처리방침 URL(운영 필수)", "● (브랜딩·법무 URL)", "○ (동일 Client ID·origins 단계적 등록)"],
            ["오픈일·DNS D-7/D-3 일정", "●", "○ (배포·스모크)"],
        ],
        {"A": 42, "B": 16, "C": 48},
    )

    ws6 = wb.create_sheet("회신승인")
    write_table(
        ws6,
        ["구분", "성명", "연락처", "확정 일자", "서명/승인"],
        [
            ["고객사(주미얀마 대사관)", "", "", "", ""],
            ["MOFA IT / DNS 담당", "", "", "", ""],
            ["운영팀(BO)", "", "", "", ""],
        ],
        {"A": 28, "B": 16, "C": 24, "D": 14, "E": 16},
    )

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
